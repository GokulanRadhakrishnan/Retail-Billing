import os
import shutil
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
    QMessageBox, QComboBox, QTableWidget, QTableWidgetItem, QHeaderView, QGroupBox, QFileDialog
)
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
import datetime
import zipfile

DATA_DIR = "data"
USER_FILE = "users.json"

class AdminWidget(QWidget):
    """Admin widget for user management, backup, and logs."""
    def __init__(self, auth_manager, parent=None):
        super().__init__(parent)
        self.auth_manager = auth_manager
        self.setLayout(QVBoxLayout())
        self.user_file = "users.json"  # AuthManager default (not used directly here but for info)
        self.build_ui()
        self.refresh_user_list()
        self.update_user_status()

    def build_ui(self) -> None:
        """Build the UI components."""
        # Login status
        status_layout = QHBoxLayout()
        self.lbl_user = QLabel()
        status_layout.addWidget(self.lbl_user)
        btn_logout = QPushButton("Logout")
        btn_logout.clicked.connect(self.handle_logout)
        status_layout.addWidget(btn_logout)
        self.layout().addLayout(status_layout)

        # New user creation section
        new_user_layout = QHBoxLayout()
        self.new_user_input = QLineEdit()
        self.new_user_input.setPlaceholderText("Enter new username")
        new_user_layout.addWidget(QLabel("New Username:"))
        new_user_layout.addWidget(self.new_user_input)

        self.pw_input = QLineEdit()
        self.pw_input.setPlaceholderText("New Password")
        self.pw_input.setEchoMode(QLineEdit.Password)
        new_user_layout.addWidget(QLabel("Password:"))
        new_user_layout.addWidget(self.pw_input)

        btn_create = QPushButton("Create User")
        btn_create.clicked.connect(self.create_user)
        new_user_layout.addWidget(btn_create)

        self.layout().addLayout(new_user_layout)

        # User management (existing users)
        manage_group = QGroupBox("User Management")
        mg_layout = QHBoxLayout()
        manage_group.setLayout(mg_layout)

        self.user_combo = QComboBox()
        mg_layout.addWidget(QLabel("Select User:"))
        mg_layout.addWidget(self.user_combo)
        self.user_combo.currentTextChanged.connect(self.handle_user_selected)

        self.role_label = QLabel("Role:")
        mg_layout.addWidget(self.role_label)

        self.pw_update_input = QLineEdit()
        self.pw_update_input.setPlaceholderText("New Password for Selected User")
        self.pw_update_input.setEchoMode(QLineEdit.Password)
        mg_layout.addWidget(self.pw_update_input)

        btn_update_pw = QPushButton("Update Password")
        btn_update_pw.clicked.connect(self.update_user_pw)
        mg_layout.addWidget(btn_update_pw)

        btn_delete = QPushButton("Remove User")
        btn_delete.clicked.connect(self.delete_user)
        mg_layout.addWidget(btn_delete)

        self.layout().addWidget(manage_group)

        # Staff activity log
        log_group = QGroupBox("Staff Activity Log")
        self.log_table = QTableWidget(0, 4)
        self.log_table.setHorizontalHeaderLabels(["File", "Entity", "Action", "User"])
        self.log_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        log_v_layout = QVBoxLayout()
        log_group.setLayout(log_v_layout)
        btn_refresh_log = QPushButton("Refresh Log")
        btn_refresh_log.clicked.connect(self.refresh_staff_log)
        log_v_layout.addWidget(btn_refresh_log)
        log_v_layout.addWidget(self.log_table)
        self.layout().addWidget(log_group)

        # Backup & Restore
        backup_group = QGroupBox("Backup / Restore Data")
        backup_layout = QHBoxLayout()
        backup_group.setLayout(backup_layout)
        btn_backup = QPushButton("Backup Now")
        btn_backup.clicked.connect(self.handle_backup)
        backup_layout.addWidget(btn_backup)
        btn_restore = QPushButton("Restore...")
        btn_restore.clicked.connect(self.handle_restore)
        backup_layout.addWidget(btn_restore)
        self.layout().addWidget(backup_group)

    def update_user_status(self) -> None:
        """Update user status label."""
        user = self.auth_manager.get_current_user()
        roles = self.auth_manager.get_user_roles(user) if user else []
        self.lbl_user.setText(f"Logged in as: {user} ({', '.join(roles)})")

    def refresh_user_list(self) -> None:
        """Refresh user list combo box."""
        self.user_combo.blockSignals(True)
        self.user_combo.clear()
        all_users = sorted(self.auth_manager.users.keys())
        self.user_combo.addItems(all_users)
        self.user_combo.blockSignals(False)
        self.handle_user_selected()

    def handle_user_selected(self) -> None:
        """Handle user selection change."""
        username = self.user_combo.currentText()
        if not username:
            self.role_label.setText("Role:")
            return
        roles = self.auth_manager.get_user_roles(username)
        self.role_label.setText(f"Role: {', '.join(roles) if roles else 'staff'}")
        self.pw_update_input.clear()  # Clear password input on user change

    def create_user(self) -> None:
        """Create a new user."""
        username = self.new_user_input.text().strip()
        password = self.pw_input.text().strip()
        if not username or not password:
            QMessageBox.warning(self, "Input Error", "Username and password required.")
            return
        valid, msg = self.auth_manager.validate_password_rules(password)
        if not valid:
            QMessageBox.warning(self, "Password Policy", msg)
            return
        try:
            self.auth_manager.add_user(username, password, roles=["staff"])
            QMessageBox.information(self, "User Created", f"Staff user '{username}' created.")
            self.refresh_user_list()
            self.new_user_input.clear()
            self.pw_input.clear()
        except Exception as e:
            QMessageBox.warning(self, "Error", str(e))

    def update_user_pw(self) -> None:
        """Update password for selected user."""
        username = self.user_combo.currentText().strip()
        password = self.pw_update_input.text().strip()
        if not username or not password:
            QMessageBox.warning(self, "Input Error", "Username and new password required.")
            return
        valid, msg = self.auth_manager.validate_password_rules(password)
        if not valid:
            QMessageBox.warning(self, "Password Policy", msg)
            return
        try:
            self.auth_manager.update_password(username, password)
            QMessageBox.information(self, "Password Updated", f"Password for '{username}' updated.")
            self.pw_update_input.clear()
        except Exception as e:
            QMessageBox.warning(self, "Error", str(e))

    def delete_user(self) -> None:
        """Delete selected user."""
        username = self.user_combo.currentText()
        if username == "admin":
            QMessageBox.warning(self, "Denied", "Cannot delete default admin user.")
            return
        confirm = QMessageBox.question(self, "Delete User",
            f"Are you sure to delete user '{username}'?", QMessageBox.Yes | QMessageBox.No)
        if confirm != QMessageBox.Yes:
            return
        try:
            self.auth_manager.delete_user(username)
            QMessageBox.information(self, "Deleted", f"User '{username}' deleted.")
            self.refresh_user_list()
        except Exception as e:
            QMessageBox.warning(self, "Error", str(e))

    def refresh_staff_log(self) -> None:
        """Refresh staff activity log table."""
        # Aggregates entries/actions from purchase/sales Excel showing staff activity.
        self.log_table.setRowCount(0)
        for fname in os.listdir(DATA_DIR):
            if not fname.endswith(".xlsx"):
                continue
            path = os.path.join(DATA_DIR, fname)
            try:
                wb = load_workbook(path, data_only=True)
                for sheetname in wb.sheetnames:
                    if sheetname not in ("Invoices", "Bills"):
                        continue
                    ws = wb[sheetname]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if sheetname == "Invoices":
                            ent = f"Invoice {row[0]}"
                            action = "Purchase"
                            user = row[9] if len(row) > 9 else ""
                        else:
                            ent = f"Bill {row[0]}"
                            action = "Sale"
                            user = row[14] if len(row) > 14 else ""
                        rowpos = self.log_table.rowCount()
                        self.log_table.insertRow(rowpos)
                        self.log_table.setItem(rowpos, 0, QTableWidgetItem(fname))
                        self.log_table.setItem(rowpos, 1, QTableWidgetItem(ent))
                        self.log_table.setItem(rowpos, 2, QTableWidgetItem(action))
                        self.log_table.setItem(rowpos, 3, QTableWidgetItem(str(user)))
            except Exception:
                # Silently skip problematic files (optional: add logging)
                continue

    def handle_backup(self) -> None:
        """Handle backup of data files."""
        backup_root_dir = QFileDialog.getExistingDirectory(self, "Select Backup Directory")
        if not backup_root_dir:
            return

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_subdir = os.path.join(backup_root_dir, f"backup_{timestamp}")
        os.makedirs(backup_subdir, exist_ok=True)

        try:
            # Build list of files to back up with explicit source and archive names
            files = []  # list of tuples (src_path, arcname)

            # Excel files in data directory
            for fname in os.listdir(DATA_DIR):
                if fname.lower().endswith(".xlsx"):
                    files.append((os.path.join(DATA_DIR, fname), fname))

            # users.json at project root
            if os.path.exists(USER_FILE):
                files.append((USER_FILE, os.path.basename(USER_FILE)))

            # SQLite databases
            if os.path.exists("purchases.db"):
                files.append(("purchases.db", "purchases.db"))
            sales_db_path = os.path.join(DATA_DIR, "sales_data.db")
            if os.path.exists(sales_db_path):
                files.append((sales_db_path, "sales_data.db"))

            # Copy to backup_subdir using arcnames
            for src, arcname in files:
                dst = os.path.join(backup_subdir, arcname)
                shutil.copy2(src, dst)

            # Zip the backup_subdir
            zip_path = os.path.join(backup_root_dir, f"backup_{timestamp}.zip")
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for _, arcname in files:
                    full_path = os.path.join(backup_subdir, arcname)
                    zipf.write(full_path, arcname=arcname)

            QMessageBox.information(
                self,
                "Backup Complete",
                f"Backup saved to:\n{backup_subdir}\n\nCompressed as:\n{zip_path}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Backup Failed", f"Backup failed:\n{e}")

    def handle_restore(self) -> None:
        """Handle restore from backup files."""
        # Let user pick a zip backup file or an Excel file for restore
        file_filter = "Backup Files (*.zip *.xlsx)"
        restore_path, _ = QFileDialog.getOpenFileName(self, "Select Backup to Restore", "", file_filter)
        if not restore_path:
            return

        try:
            # If zip file, unzip contents to a temp folder and restore files
            if restore_path.lower().endswith(".zip"):
                import tempfile
                with tempfile.TemporaryDirectory() as tempdir:
                    with zipfile.ZipFile(restore_path, 'r') as zf:
                        zf.extractall(tempdir)

                    # Restore extracted files to their appropriate locations
                    for fname in os.listdir(tempdir):
                        src = os.path.join(tempdir, fname)
                        # Decide destination based on filename
                        if fname.lower().endswith('.xlsx'):
                            dst = os.path.join(DATA_DIR, fname)
                        elif fname == os.path.basename(USER_FILE) or fname == "purchases.db":
                            dst = fname  # project root
                        elif fname == "sales_data.db":
                            dst = os.path.join(DATA_DIR, "sales_data.db")
                        else:
                            # Skip unknown files
                            continue

                        confirm = QMessageBox.question(
                            self,
                            "Confirm Restore",
                            f"Restore file '{fname}' and overwrite existing data if it exists?",
                            QMessageBox.Yes | QMessageBox.No,
                            QMessageBox.No
                        )
                        if confirm == QMessageBox.Yes:
                            # Ensure directory exists for DATA_DIR destinations
                            os.makedirs(os.path.dirname(dst) or '.', exist_ok=True)
                            shutil.copy2(src, dst)

            elif restore_path.lower().endswith(".xlsx"):
                # Single Excel file restore
                fname = os.path.basename(restore_path)
                dst = os.path.join(DATA_DIR, fname)
                confirm = QMessageBox.question(
                    self,
                    "Confirm Restore",
                    f"Restore file '{fname}' and overwrite existing data if it exists?",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )
                if confirm == QMessageBox.Yes:
                    shutil.copy2(restore_path, dst)
            else:
                QMessageBox.warning(self, "Invalid File", "Please select a valid backup (.zip) or Excel (.xlsx) file.")
                return

            QMessageBox.information(self, "Restore Complete", "Data restore completed successfully.")

        except Exception as e:
            QMessageBox.critical(self, "Restore Failed", f"Restore failed:\n{e}")

    def handle_logout(self) -> None:
        """Logout current user."""
        self.auth_manager.logout()
        QMessageBox.information(self, "Logged Out", "You have been logged out.")
        self.update_user_status()
