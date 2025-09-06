import os
import sqlite3
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QMessageBox,
    QComboBox, QTableWidget, QTableWidgetItem, QHeaderView, QDateEdit, QGroupBox,
    QListWidget, QAbstractItemView, QSplitter, QFormLayout, QCompleter
)
from PyQt5.QtCore import Qt, QDate, QEvent
from openpyxl import Workbook, load_workbook
from datetime import datetime

PURCHASE_FILE_DIR = "data"
UNIT_OPTIONS = ["g", "kg", "ml", "l", "pcs", "bag"]
GST_OPTIONS = ["NIL", "0", "5", "12", "18", "28"]
CATEGORY_OPTIONS = ["Seeds", "Pesticide", "Fertilizer"]
SQLITE_DB_PATH_PURCHASE = "purchases.db"  # Use the same DB as the rest of this module

DB_FILE = "purchases.db"
INV_DB_PATH = "data/sales_data.db"

def ensure_inventory_db() -> None:
    """Ensure inventory DB and table exist."""
    try:
        os.makedirs(os.path.dirname(INV_DB_PATH), exist_ok=True)
        conn = sqlite3.connect(INV_DB_PATH)
        c = conn.cursor()
        c.execute(
            '''CREATE TABLE IF NOT EXISTS purchase_stock (
                   product_name TEXT PRIMARY KEY,
                   quantity REAL
               )'''
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"WARNING: ensure_inventory_db failed: {e}")

def inventory_add(product_name: str, qty: float) -> None:
    """Add quantity to inventory for a product."""
    ensure_inventory_db()
    try:
        conn = sqlite3.connect(INV_DB_PATH)
        c = conn.cursor()
        c.execute("SELECT quantity FROM purchase_stock WHERE product_name=?", (product_name,))
        row = c.fetchone()
        if row:
            new_qty = float(row[0] or 0) + float(qty or 0)
            c.execute("UPDATE purchase_stock SET quantity=? WHERE product_name=?", (new_qty, product_name))
        else:
            c.execute("INSERT INTO purchase_stock(product_name, quantity) VALUES (?, ?)", (product_name, float(qty or 0)))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"WARNING: inventory_add failed: {e}")

def inventory_subtract(product_name: str, qty: float) -> None:
    """Subtract quantity from inventory for a product."""
    ensure_inventory_db()
    try:
        conn = sqlite3.connect(INV_DB_PATH)
        c = conn.cursor()
        c.execute("SELECT quantity FROM purchase_stock WHERE product_name=?", (product_name,))
        row = c.fetchone()
        if row:
            new_qty = max(0.0, float(row[0] or 0) - float(qty or 0))
            c.execute("UPDATE purchase_stock SET quantity=? WHERE product_name=?", (new_qty, product_name))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"WARNING: inventory_subtract failed: {e}")

def init_db() -> None:
    """Initialize SQLite DB and table if not exists."""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS purchases (
            invoice_no TEXT,
            date TEXT,
            vendor TEXT,
            product_name TEXT,
            qty REAL,
            unit TEXT,
            mrp REAL,
            gst TEXT,
            expiry TEXT,
            category TEXT,
            entry_by TEXT,
            PRIMARY KEY (invoice_no, product_name, unit, mrp, gst, expiry)
        )
    ''')
    conn.commit()
    conn.close()

def insert_purchase(purchase_tuple: tuple) -> None:
    """Insert or replace purchase record into SQLite."""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        INSERT OR REPLACE INTO purchases (
            invoice_no, date, vendor, product_name, qty, unit, mrp, gst, expiry, category, entry_by
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', purchase_tuple)
    conn.commit()
    conn.close()

def delete_invoice_from_db(invoice_no: str) -> None:
    """Delete all records of an invoice from SQLite"""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("DELETE FROM purchases WHERE invoice_no=?", (invoice_no,))
    conn.commit()
    conn.close()


def financial_year_for_date(date: QDate) -> str:
    """Return financial year string for a given QDate."""
    month = date.month()
    year = date.year()
    if month >= 4:
        return f"{year}-{year+1}"
    else:
        return f"{year-1}-{year}"


def purchase_excel_path(date: QDate) -> str:
    """Return purchase Excel file path for a given QDate."""
    fy_start = date.year() if date.month() >= 4 else date.year() - 1
    fy_end = fy_start + 1
    fname = f"Purchase_{fy_start}-{fy_end}.xlsx"
    return os.path.join(PURCHASE_FILE_DIR, fname)


def carry_forward_purchase_fy_stock(now_date: QDate, purchase_excel_path_func) -> None:
    """Carry forward stock from previous FY to new FY purchase file."""
    if now_date.month() < 4:
        # Do not run carry forward before April
        return

    cur_fy_start = now_date.year() if now_date.month() >= 4 else now_date.year() - 1
    prev_fy_start = cur_fy_start - 1
    prev_fy_end = cur_fy_start

    prev_path = purchase_excel_path_func(QDate(prev_fy_end, 3, 31))
    new_path = purchase_excel_path_func(QDate(cur_fy_start, 4, 1))

    if not os.path.exists(prev_path):
        print(f"Previous purchase file for FY {prev_fy_start}-{prev_fy_end} not found: {prev_path}")
        return

    # Load previous FY workbook and sheet
    prev_wb = load_workbook(prev_path)
    if "Invoices" not in prev_wb.sheetnames:
        print(f"No 'Invoices' sheet in previous FY purchase file: {prev_path}")
        return
    prev_ws = prev_wb["Invoices"]
    headers = [cell.value for cell in prev_ws[1]]

    # Collect all products from previous FY with qty > 0, consolidating by product + unit + mrp + gst + expiry + category
    product_stock_map = {}
    for row in prev_ws.iter_rows(min_row=2, values_only=True):
        pname = row[3]
        qty = row[4]
        unit = row[5]
        mrp = row[6]
        gst = row[7]
        expiry = row[8]
        category = row[9] if len(row) > 9 else ""
        # Only consider products with positive qty
        if pname and qty and qty > 0:
            key = (pname, unit, mrp, gst, expiry, category)
            product_stock_map[key] = product_stock_map.get(key, 0) + qty

    if not product_stock_map:
        print("No stock to carry forward from previous FY.")
        return

    # If new FY file doesn't exist, create it with same structure
    if not os.path.exists(new_path):
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = "Invoices"
        new_ws.append(headers)
        for (pname, unit, mrp, gst, expiry, category), qty in product_stock_map.items():
            # Compose a row with empty invoice info (e.g. Invoice No, Date, Vendor) at front,
            # as this is "opening stock" – you can set Invoice No as "Opening Stock", Date as 01-04-YYYY, Vendor blank
            new_row = [
                "Opening Stock",  # Invoice No
                f"01-04-{cur_fy_start}",  # Date string
                "",  # Vendor
                pname,
                qty,
                unit,
                mrp,
                gst,
                expiry if expiry else "",
                category if category else "",
                "carry_forward"
            ]
            new_ws.append(new_row)
        new_wb.save(new_path)
        print(f"Created new FY purchase file and carried forward stock: {new_path}")
    else:
        # If new FY file exists, append only those products that do not already exist as Opening Stock
        wb = load_workbook(new_path)
        if "Invoices" not in wb.sheetnames:
            print(f"No 'Invoices' sheet in new FY purchase file: {new_path}")
            return
        ws = wb["Invoices"]

        # Find existing Opening Stock rows to prevent duplicate carry forward
        existing_opening = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "Opening Stock":
                key = (row[3], row[5], row[6], row[7], row[8], row[9] if len(row) > 9 else "")
                existing_opening.add(key)

        rows_added = 0
        for key, qty in product_stock_map.items():
            if key not in existing_opening:
                pname, unit, mrp, gst, expiry, category = key
                new_row = [
                    "Opening Stock",
                    f"01-04-{cur_fy_start}",
                    "",
                    pname,
                    qty,
                    unit,
                    mrp,
                    gst,
                    expiry if expiry else "",
                    category if category else "",
                    "carry_forward"
                ]
                ws.append(new_row)
                rows_added += 1
        if rows_added:
            wb.save(new_path)
            print(f"Appended {rows_added} opening stock items to existing FY purchase file: {new_path}")
        else:
            print(f"No new opening stock entries to add in {new_path}.")
            
def get_product_category_from_db(product_name: str) -> str:
    """Look up the latest category for a product from the purchases table."""
    db_path = DB_FILE  # unify to purchases.db
    if not os.path.exists(db_path):
        return ""
    try:
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        c.execute(
            """
            SELECT category FROM purchases
            WHERE product_name = ?
            ORDER BY rowid DESC LIMIT 1
            """,
            (product_name,)
        )
        row = c.fetchone()
        conn.close()
        return row[0] if row else ""
    except Exception:
        return ""


class PurchaseWidget(QWidget):
    """Widget for purchase entry and inventory management."""
    def __init__(self, auth_manager=None, parent=None):
        super().__init__(parent)
        init_db()  # Ensure DB initialized

        self.auth_manager = auth_manager
        self.current_invoice_products = []  # Each product: [Product, Qty, Unit, MRP, GST, Expiry, Category]
        self.current_invoice_no = None
        self.current_invoice_date = QDate.currentDate()
        self.current_invoice_vendor = ""
        self.is_editing = False  # Flag to distinguish new vs edit mode

        self.excel_path = purchase_excel_path(QDate.currentDate())
        self.ensure_excel_structure(self.excel_path)

        self.setLayout(QVBoxLayout())
        self.build_ui()
        self.load_invoice_list()

    def build_ui(self) -> None:
        """Build the UI components."""
        splitter = QSplitter(Qt.Horizontal)
        self.layout().addWidget(splitter)

        # Left side: Invoice list and search
        left_panel = QWidget()
        left_panel.setLayout(QVBoxLayout())
        splitter.addWidget(left_panel)

        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search invoice by Invoice No or Vendor...")
        self.search_input.textChanged.connect(self.handle_search_invoices)
        search_layout.addWidget(QLabel("Search:"))
        search_layout.addWidget(self.search_input)
        left_panel.layout().addLayout(search_layout)

        self.invoice_listwidget = QListWidget()
        self.invoice_listwidget.setSelectionMode(QAbstractItemView.SingleSelection)
        self.invoice_listwidget.itemSelectionChanged.connect(self.load_selected_invoice)
        left_panel.layout().addWidget(self.invoice_listwidget)

        btn_del_invoice = QPushButton("Delete Selected Invoice")
        btn_del_invoice.clicked.connect(self.handle_delete_invoice)
        left_panel.layout().addWidget(btn_del_invoice)

        btn_clear_selection = QPushButton("Clear Selection / New Invoice")
        btn_clear_selection.clicked.connect(self.clear_all_fields)
        left_panel.layout().addWidget(btn_clear_selection)

        # Right side: Invoice details and product entry
        right_panel = QWidget()
        right_panel.setLayout(QVBoxLayout())
        splitter.addWidget(right_panel)

        self.build_invoice_details_section(right_panel)
        self.build_product_entry_section(right_panel)
        self.build_products_table_section(right_panel)
        self.build_save_controls_section(right_panel)
        self.setup_enter_key_navigation()

        splitter.setStretchFactor(0, 2)
        splitter.setStretchFactor(1, 5)

    def build_invoice_details_section(self, parent) -> None:
        """Build invoice details section."""
        invoice_group = QGroupBox("Invoice Details")
        parent.layout().addWidget(invoice_group)
        form = QFormLayout()
        invoice_group.setLayout(form)

        self.invoice_no = QLineEdit()
        form.addRow(QLabel("Invoice Number:"), self.invoice_no)

        self.date_edit = QDateEdit(QDate.currentDate())
        self.date_edit.setDisplayFormat("dd-MM-yyyy")
        self.date_edit.setCalendarPopup(True)
        form.addRow(QLabel("Date:"), self.date_edit)

        self.vendor_name = QLineEdit()
        form.addRow(QLabel("Vendor Name:"), self.vendor_name)

    def build_product_entry_section(self, parent) -> None:
        """Build product entry section."""
        prod_group = QGroupBox("Add/Edit Product")
        parent.layout().addWidget(prod_group)
        layout = QHBoxLayout()
        prod_group.setLayout(layout)

        self.product_name = QLineEdit()
        self.product_name.setPlaceholderText("Product Name")
        # Attach completer with existing product names
        try:
            completer = QCompleter(self.get_distinct_product_names())
            completer.setCaseSensitivity(Qt.CaseInsensitive)
            self.product_name.setCompleter(completer)
        except Exception:
            pass
        # Auto-fill category on product commit
        self.product_name.editingFinished.connect(self.autofill_category_for_product)
        layout.addWidget(QLabel("Product:"))
        layout.addWidget(self.product_name)

        self.qty = QLineEdit()
        self.qty.setPlaceholderText("Quantity")
        layout.addWidget(QLabel("Qty:"))
        layout.addWidget(self.qty)

        self.unit = QComboBox()
        self.unit.addItems(UNIT_OPTIONS)
        layout.addWidget(QLabel("Unit:"))
        layout.addWidget(self.unit)

        self.mrp = QLineEdit()
        self.mrp.setPlaceholderText("MRP incl GST")
        layout.addWidget(QLabel("MRP:"))
        layout.addWidget(self.mrp)

        self.gst = QComboBox()
        self.gst.addItems(GST_OPTIONS)
        layout.addWidget(QLabel("GST %:"))
        layout.addWidget(self.gst)

        self.expiry = QDateEdit()
        self.expiry.setMinimumDate(QDate(2000, 1, 1))
        self.expiry.setSpecialValueText("No Expiry")
        self.expiry.setDate(self.expiry.minimumDate())
        self.expiry.setCalendarPopup(True)
        self.expiry.setDisplayFormat("dd-MM-yyyy")
        layout.addWidget(QLabel("Expiry Date:"))
        layout.addWidget(self.expiry)

        self.category = QComboBox()
        self.category.addItems(CATEGORY_OPTIONS)
        layout.addWidget(QLabel("Category:"))
        layout.addWidget(self.category)

        self.btn_add_edit_product = QPushButton("Add Product")
        self.btn_add_edit_product.clicked.connect(self.handle_add_edit_product)
        layout.addWidget(self.btn_add_edit_product)

        self.btn_cancel_edit = QPushButton("Cancel Edit")
        self.btn_cancel_edit.clicked.connect(self.cancel_product_edit)
        self.btn_cancel_edit.setVisible(False)  # Only visible during product edit mode
        layout.addWidget(self.btn_cancel_edit)

    def build_products_table_section(self, parent) -> None:
        """Build products table section."""
        table_group = QGroupBox("Products in Current Invoice")
        parent.layout().addWidget(table_group)
        self.product_table = QTableWidget(0, 7)
        self.product_table.setHorizontalHeaderLabels(
            ["Product", "Qty", "Unit", "MRP", "GST %", "Expiry", "Category"]
        )
        self.product_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.product_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.product_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.product_table.doubleClicked.connect(self.edit_selected_product)
        layout = QVBoxLayout()
        table_group.setLayout(layout)
        layout.addWidget(self.product_table)

        btn_layout = QHBoxLayout()
        layout.addLayout(btn_layout)

        btn_remove_product = QPushButton("Remove Selected Product")
        btn_remove_product.clicked.connect(self.handle_remove_product)
        btn_layout.addWidget(btn_remove_product)

        btn_edit_product = QPushButton("Edit Selected Product")
        btn_edit_product.clicked.connect(self.edit_selected_product)
        btn_layout.addWidget(btn_edit_product)

    def build_save_controls_section(self, parent) -> None:
        """Build save controls section."""
        btn_layout = QHBoxLayout()
        parent.layout().addLayout(btn_layout)

        self.btn_save_invoice = QPushButton("Save Invoice")
        self.btn_save_invoice.clicked.connect(self.handle_save_invoice)
        btn_layout.addWidget(self.btn_save_invoice)

        self.btn_clear_all = QPushButton("Clear All Fields")
        self.btn_clear_all.clicked.connect(self.clear_all_fields)
        btn_layout.addWidget(self.btn_clear_all)

    def ensure_excel_structure(self, path: str) -> None:
        """Ensure Excel file exists with required sheets and headers."""
        os.makedirs(PURCHASE_FILE_DIR, exist_ok=True)
        if not os.path.exists(path):
            wb = Workbook()
            ws_inv = wb.active
            ws_inv.title = "Invoices"
            ws_inv.append([
                "Invoice No", "Date", "Vendor", "Product", "Qty",
                "Unit", "MRP", "GST %", "Expiry", "Category", "Entry By"
            ])
            ws_vendor = wb.create_sheet("VendorWise")
            ws_vendor.append([
                "Vendor", "Invoice No", "Date", "Product", "Qty",
                "Unit", "MRP", "GST %", "Expiry", "Category", "Entry By"
            ])
            wb.save(path)

    def load_invoice_list(self) -> None:
        """Load invoice list from Excel and DB."""
        self.invoice_listwidget.clear()
        invoices = set()
        vendor_map = {}

        # Load from Excel
        if os.path.exists(self.excel_path):
            try:
                wb = load_workbook(self.excel_path, data_only=True)
                sheet = wb["Invoices"]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    invoice_no = row[0]
                    vendor = row[2]
                    if invoice_no and invoice_no not in invoices:
                        invoices.add(invoice_no)
                        vendor_map[invoice_no] = vendor
            except Exception as e:
                QMessageBox.warning(self, "Load Error", f"Failed to load invoices from Excel: {str(e)}")

        # Load from SQLite
        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT invoice_no, vendor FROM purchases")
            for invoice_no, vendor in cursor.fetchall():
                if invoice_no and invoice_no not in invoices:
                    invoices.add(invoice_no)
                    vendor_map[invoice_no] = vendor
            conn.close()
        except Exception as e:
            QMessageBox.warning(self, "Load Error", f"Failed to load invoices from database: {str(e)}")

        # Sort and display as "InvoiceNo | Vendor"
        sorted_invoices = sorted(invoices)
        for inv in sorted_invoices:
            display_text = f"{inv} | {vendor_map.get(inv, '')}"
            self.invoice_listwidget.addItem(display_text)

    def handle_search_invoices(self, text: str) -> None:
        """Handle invoice search input."""
        text = text.strip().lower()
        for i in range(self.invoice_listwidget.count()):
            item = self.invoice_listwidget.item(i)
            item.setHidden(text not in item.text().lower())

    def load_selected_invoice(self) -> None:
        """Load selected invoice details."""
        selected_items = self.invoice_listwidget.selectedItems()
        if not selected_items:
            return
        selected_text = selected_items[0].text()
        invoice_no = selected_text.split("|")[0].strip()
        if not invoice_no:
            return
        self.load_invoice(invoice_no)

    def load_invoice(self, invoice_no: str) -> None:
        """Load invoice details from DB or Excel."""
        # Try loading from SQLite first
        try:
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT date, vendor, product_name, qty, unit, mrp, gst, expiry, category
                FROM purchases WHERE invoice_no=?
            """, (invoice_no,))
            rows = cursor.fetchall()
            conn.close()

            if rows:
                # Use SQLite data
                invoice_date_str = rows[0][0]
                invoice_vendor = rows[0][1]
                products = []
                for row in rows:
                    products.append({
                        "Product": row[2],
                        "Qty": row[3],
                        "Unit": row[4],
                        "MRP": row[5],
                        "GST": row[6],
                        "Expiry": row[7] or "",
                        "Category": row[8] or ""
                    })

                self.current_invoice_no = invoice_no
                self.current_invoice_vendor = invoice_vendor
                self.current_invoice_date = QDate.fromString(invoice_date_str, "dd-MM-yyyy")
                self.is_editing = True

                self.invoice_no.setText(invoice_no)
                self.date_edit.setDate(self.current_invoice_date)
                self.vendor_name.setText(invoice_vendor)

                self.current_invoice_products = []
                for p in products:
                    self.current_invoice_products.append([
                        p["Product"], p["Qty"], p["Unit"], p["MRP"], p["GST"], p["Expiry"], p["Category"]
                    ])
                self.refresh_product_table()
                self.set_product_entry_mode(add=True)
                return  # loaded from DB, done
        except Exception as e:
            print(f"Warning: Could not load invoice from DB: {e}")

        # If not found in DB or DB error, fallback to Excel
        path = self.excel_path
        if not os.path.exists(path):
            QMessageBox.warning(self, "File Missing", "Purchase Excel file not found for current financial year.")
            return

        try:
            wb = load_workbook(path)
            sheet = wb["Invoices"]
            products = []
            invoice_date = None
            invoice_vendor = None
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == invoice_no:
                    invoice_date = row[1]
                    invoice_vendor = row[2]
                    products.append({
                        "Product": row[3],
                        "Qty": row[4],
                        "Unit": row[5],
                        "MRP": row[6],
                        "GST": row[7],
                        "Expiry": row[8] or "",
                        "Category": row[9] or ""
                    })
            if not products:
                QMessageBox.information(self, "Not Found", f"No products found for invoice {invoice_no}.")
                return

            self.current_invoice_no = invoice_no
            self.current_invoice_vendor = invoice_vendor
            self.current_invoice_date = QDate.fromString(invoice_date, "dd-MM-yyyy")
            self.is_editing = True

            self.invoice_no.setText(invoice_no)
            self.date_edit.setDate(self.current_invoice_date)
            self.vendor_name.setText(invoice_vendor)

            self.current_invoice_products = []
            for p in products:
                self.current_invoice_products.append([
                    p["Product"], p["Qty"], p["Unit"], p["MRP"], p["GST"], p["Expiry"], p["Category"]
                ])
            self.refresh_product_table()
            self.set_product_entry_mode(add=True)

        except Exception as e:
            QMessageBox.warning(self, "Load Error", f"Unable to load invoice: {str(e)}")

    def handle_add_edit_product(self) -> None:
        """Add or edit product entry in invoice."""
        pname = self.product_name.text().strip()
        qty = self.qty.text().strip()
        unit = self.unit.currentText()
        mrp = self.mrp.text().strip()
        gst = self.gst.currentText()
        expiry_date = self.expiry.date()
        if expiry_date == self.expiry.minimumDate():
            expiry = ""
        else:
            expiry = expiry_date.toString("dd-MM-yyyy")
        category = self.category.currentText()

        if not pname or not qty or not mrp:
            QMessageBox.warning(self, "Input Error", "Product Name, Quantity, and MRP are required.")
            return
        try:
            qty_val = float(qty)
            mrp_val = float(mrp)
            if qty_val <= 0 or mrp_val < 0:
                raise ValueError()
        except ValueError:
            QMessageBox.warning(self, "Validation Error", "Quantity must be positive and MRP cannot be negative.")
            return

        product_data = [pname, qty_val, unit, mrp_val, gst, expiry, category]

        if self.btn_add_edit_product.text() == "Add Product":
            self.current_invoice_products.append(product_data)
            self.refresh_product_table()
            self.clear_product_entry_fields()
            self.product_name.setFocus()
        else:  # Editing existing product
            self.current_edit_product_row = getattr(self, "current_edit_product_row", None)
            if self.current_edit_product_row is not None:
                self.current_invoice_products[self.current_edit_product_row] = product_data
                self.refresh_product_table()
                self.clear_product_entry_fields()
                self.set_product_entry_mode(add=True)
                self.product_name.setFocus()

    def refresh_product_table(self) -> None:
        """Refresh products table display."""
        self.product_table.setRowCount(len(self.current_invoice_products))
        for row_idx, prod in enumerate(self.current_invoice_products):
            for col_idx, val in enumerate(prod):
                self.product_table.setItem(row_idx, col_idx, QTableWidgetItem(str(val)))

    def clear_product_entry_fields(self) -> None:
        """Clear product entry fields."""
        self.product_name.setText("")
        self.qty.setText("")
        self.mrp.setText("")
        self.gst.setCurrentIndex(0)
        self.unit.setCurrentIndex(0)
        self.expiry.setDate(self.expiry.minimumDate())
        self.category.setCurrentIndex(0)

    def handle_remove_product(self) -> None:
        """Remove selected product from invoice."""
        selected = self.product_table.selectionModel().selectedRows()
        if not selected:
            QMessageBox.information(self, "Remove Product", "Select a product row to remove.")
            return
        idx = selected[0].row()
        self.current_invoice_products.pop(idx)
        self.refresh_product_table()

    def edit_selected_product(self) -> None:
        """Edit selected product in invoice."""
        selected = self.product_table.selectionModel().selectedRows()
        if not selected:
            QMessageBox.information(self, "Edit Product", "Select a product row to edit.")
            return
        idx = selected[0].row()
        prod = self.current_invoice_products[idx]
        self.product_name.setText(str(prod[0]))
        self.qty.setText(str(prod[1]))
        self.unit.setCurrentText(prod[2])
        self.mrp.setText(str(prod[3]))
        gst_index = self.gst.findText(prod[4])
        self.gst.setCurrentIndex(gst_index if gst_index >= 0 else 0)
        if prod[5]:
            try:
                dt = QDate.fromString(prod[5], "dd-MM-yyyy")
                if dt.isValid():
                    self.expiry.setDate(dt)
                else:
                    self.expiry.setDate(self.expiry.minimumDate())
            except Exception:
                self.expiry.setDate(self.expiry.minimumDate())
        else:
            self.expiry.setDate(self.expiry.minimumDate())
        cat_index = self.category.findText(prod[6])
        self.category.setCurrentIndex(cat_index if cat_index >= 0 else 0)

        self.set_product_entry_mode(add=False)
        self.current_edit_product_row = idx

    def set_product_entry_mode(self, add=True) -> None:
        """Set product entry mode (add/edit)."""
        if add:
            self.btn_add_edit_product.setText("Add Product")
            self.btn_cancel_edit.setVisible(False)
            self.current_edit_product_row = None
        else:
            self.btn_add_edit_product.setText("Save Product")
            self.btn_cancel_edit.setVisible(True)

    def setup_enter_key_navigation(self) -> None:
        """Setup Enter key navigation for fields."""
        self.invoice_no.returnPressed.connect(self.date_edit.setFocus)
        self.vendor_name.returnPressed.connect(self.product_name.setFocus)
        self.product_name.returnPressed.connect(self.qty.setFocus)
        self.qty.returnPressed.connect(self.unit.setFocus)
        self.mrp.returnPressed.connect(self.gst.setFocus)

        self.date_edit.installEventFilter(self)
        self.unit.installEventFilter(self)
        self.gst.installEventFilter(self)
        self.expiry.installEventFilter(self)

    def eventFilter(self, source, event):
        """Handle Enter key navigation for widgets."""
        if event.type() == QEvent.KeyPress and event.key() == Qt.Key_Return:
            if source == self.date_edit:
                self.vendor_name.setFocus()
                return True
            elif source == self.unit:
                self.mrp.setFocus()
                return True
            elif source == self.gst:
                self.expiry.setFocus()
                return True
            elif source == self.expiry:
                self.handle_add_edit_product()
                return True
        return super().eventFilter(source, event)

    def cancel_product_edit(self) -> None:
        """Cancel product edit mode."""
        self.clear_product_entry_fields()
        self.set_product_entry_mode(add=True)

    def get_distinct_product_names(self) -> list:
        """Get distinct product names from DB."""
        names = []
        try:
            conn = sqlite3.connect(DB_FILE)
            cur = conn.cursor()
            cur.execute("SELECT DISTINCT product_name FROM purchases ORDER BY product_name ASC")
            names = [row[0] for row in cur.fetchall() if row and row[0]]
            conn.close()
        except Exception:
            pass
        return names

    def normalize_category(self, category: str) -> str | None:
        """Normalize category string to standard values."""
        if not category:
            return None
        s = str(category).strip().lower()
        if "seed" in s:
            return "Seeds"
        if "pesticid" in s:
            return "Pesticide"
        if "fertil" in s:
            return "Fertilizer"
        return None

    def autofill_category_for_product(self) -> None:
        """Autofill category field for product."""
        pname = self.product_name.text().strip()
        if not pname:
            return
        # Look up last category from DB
        try:
            conn = sqlite3.connect(DB_FILE)
            cur = conn.cursor()
            cur.execute(
                "SELECT category FROM purchases WHERE product_name=? ORDER BY rowid DESC LIMIT 1",
                (pname,)
            )
            row = cur.fetchone()
            conn.close()
            if row and row[0]:
                cat = self.normalize_category(row[0])
                if cat and cat in CATEGORY_OPTIONS:
                    self.category.setCurrentText(cat)
                    return
        except Exception:
            pass
        # Heuristic fallback based on name
        cat = self.normalize_category(pname)
        if cat and cat in CATEGORY_OPTIONS:
            self.category.setCurrentText(cat)

    def validate_invoice_fields(self) -> bool:
        """Validate invoice fields before saving."""
        inv_no = self.invoice_no.text().strip()
        vendor = self.vendor_name.text().strip()
        if not inv_no:
            QMessageBox.warning(self, "Validation", "Invoice Number is required.")
            return False
        if not vendor:
            QMessageBox.warning(self, "Validation", "Vendor Name is required.")
            return False
        if len(self.current_invoice_products) == 0:
            QMessageBox.warning(self, "Validation", "Add at least one product to save invoice.")
            return False
        return True

    def handle_save_invoice(self) -> None:
        """Save invoice and update inventory."""
        if not self.validate_invoice_fields():
            return

        inv_no = self.invoice_no.text().strip()
        vendor = self.vendor_name.text().strip()
        date_obj = self.date_edit.date()
        date_str = date_obj.toString("dd-MM-yyyy")
        entry_by = self.auth_manager.get_current_user() if self.auth_manager else "unknown"

        self.excel_path = purchase_excel_path(date_obj)
        self.ensure_excel_structure(self.excel_path)

        try:
            wb = load_workbook(self.excel_path)
            inv_ws = wb["Invoices"]
            vendor_ws = wb["VendorWise"]

            # Delete existing invoice rows if editing and adjust inventory
            if self.is_editing and self.current_invoice_no:
                # Capture previous products to reverse inventory
                try:
                    conn_prev = sqlite3.connect(DB_FILE)
                    cur_prev = conn_prev.cursor()
                    cur_prev.execute("SELECT product_name, qty FROM purchases WHERE invoice_no=?", (self.current_invoice_no,))
                    prev_products = cur_prev.fetchall()
                    conn_prev.close()
                except Exception:
                    prev_products = []

                rows_to_delete = [row[0].row for row in inv_ws.iter_rows(min_row=2) if row[0].value == self.current_invoice_no]
                for r in reversed(rows_to_delete):
                    inv_ws.delete_rows(r)

                rows_to_delete_v = [row[0].row for row in vendor_ws.iter_rows(min_row=2) if row[1].value == self.current_invoice_no]
                for r in reversed(rows_to_delete_v):
                    vendor_ws.delete_rows(r)

                # Delete from DB as well
                delete_invoice_from_db(self.current_invoice_no)

                # Adjust inventory by subtracting previous quantities
                for pname, qty_prev in prev_products:
                    try:
                        inventory_subtract(pname, float(qty_prev or 0))
                    except Exception:
                        pass

            # Append current invoice products to Excel and DB
            for prod in self.current_invoice_products:
                pname, qty, unit, mrp, gst, expiry, category = prod
                inv_ws.append([inv_no, date_str, vendor, pname, qty, unit, mrp, gst, expiry, category, entry_by])
                vendor_ws.append([vendor, inv_no, date_str, pname, qty, unit, mrp, gst, expiry, category, entry_by])

                # Also write to DB
                purchase_tuple = (
                    inv_no,                # Invoice Number (str)
                    date_str,              # Date string (str)
                    vendor,                # Vendor name (str)
                    pname,                 # Product name (str)
                    qty,                   # Quantity (float)
                    unit,                  # Unit (str)
                    mrp,                   # MRP (float)
                    gst,                   # GST (str)
                    expiry,                # Expiry date string (str)
                    category,              # Category (str)
                    entry_by               # User who entered the record (str)
                )
                try:
                    insert_purchase(purchase_tuple)
                except Exception as e:
                    print(f"WARNING: Failed to insert purchase into database: {e}")

            wb.save(self.excel_path)

            # Update inventory for new products
            try:
                for prod in self.current_invoice_products:
                    pname, qty, unit, mrp, gst, expiry, category = prod
                    inventory_add(pname, float(qty or 0))
            except Exception as e:
                print(f"WARNING: Failed to update inventory: {e}")

            QMessageBox.information(self, "Saved", f"Invoice '{inv_no}' saved successfully.")

            self.current_invoice_no = inv_no
            self.is_editing = True
            self.load_invoice_list()

        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Failed to save invoice:\n{str(e)}")

    def clear_all_fields(self) -> None:
        """Reset all invoice fields and clear the product list."""
        self.invoice_no.setText("")
        self.vendor_name.setText("")
        self.date_edit.setDate(QDate.currentDate())
        self.current_invoice_products = []
        self.refresh_product_table()
        self.clear_product_entry_fields()
        self.set_product_entry_mode(add=True)
        self.invoice_listwidget.clearSelection()
        self.is_editing = False
        self.current_invoice_no = None

    def handle_delete_invoice(self) -> None:
        """Delete selected invoice and adjust inventory."""
        if not self.auth_manager or not self.auth_manager.has_role("admin"):
           QMessageBox.warning(self, "Access Denied", "Only admin users can delete invoices.")
           return

        selected_items = self.invoice_listwidget.selectedItems()
        if not selected_items:
           QMessageBox.information(self, "Delete Invoice", "Select an invoice to delete.")
           return

        selected_text = selected_items[0].text()
        invoice_no = selected_text.split("|")[0].strip()
        if not invoice_no:
           QMessageBox.warning(self, "Delete Invoice", "Invalid invoice selected.")
           return

        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Are you sure you want to delete invoice '{invoice_no}' and all its products?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            try:
               # Capture products to adjust inventory
               try:
                   conn_prev = sqlite3.connect(DB_FILE)
                   cursor_prev = conn_prev.cursor()
                   cursor_prev.execute("SELECT product_name, qty FROM purchases WHERE invoice_no=?", (invoice_no,))
                   prev_products = cursor_prev.fetchall()
                   conn_prev.close()
               except Exception:
                   prev_products = []

               # Delete from Excel
               if os.path.exists(self.excel_path):
                   wb = load_workbook(self.excel_path)
                   inv_ws = wb["Invoices"]
                   vendor_ws = wb["VendorWise"]

                   # Delete rows matching invoice_no in Invoices sheet
                   rows_to_delete = [row[0].row for row in inv_ws.iter_rows(min_row=2) if row[0].value == invoice_no]
                   for r in reversed(rows_to_delete):
                      inv_ws.delete_rows(r)

                   # Delete rows matching invoice_no in VendorWise sheet
                   rows_to_delete_v = [row[0].row for row in vendor_ws.iter_rows(min_row=2) if row[1].value == invoice_no]
                   for r in reversed(rows_to_delete_v):
                       vendor_ws.delete_rows(r)

                   wb.save(self.excel_path)

               # Delete from SQLite DB
               delete_invoice_from_db(invoice_no)

               # Adjust inventory by subtracting deleted invoice quantities
               for pname, qty_prev in prev_products:
                   try:
                       inventory_subtract(pname, float(qty_prev or 0))
                   except Exception:
                       pass

               QMessageBox.information(self, "Deleted", f"Invoice '{invoice_no}' deleted successfully.")
               self.load_invoice_list()
               self.clear_all_fields()
               self.is_editing = False
               self.current_invoice_no = None

            except Exception as e:
                QMessageBox.critical(self, "Delete Error", f"Failed to delete invoice:\n{str(e)}")
