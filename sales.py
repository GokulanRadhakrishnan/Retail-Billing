import os
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QMessageBox,
    QComboBox, QTableWidget, QTableWidgetItem, QHeaderView, QCompleter, QSizePolicy, QDialog, QAbstractItemView
)
from PyQt5.QtCore import Qt, QDate, QStringListModel, QRegularExpression
from PyQt5.QtGui import QRegularExpressionValidator
from openpyxl import Workbook, load_workbook
from datetime import datetime
import sqlite3
from purchases import carry_forward_purchase_fy_stock, purchase_excel_path, get_product_category_from_db
from utils import update_customer_data_file, ensure_customer_data_file, add_or_update_customer, CUSTOMER_DATA_FILE, log_error, log_info
from printing import show_bill_print_preview

SALES_FILE_DIR = "data"
PAYMENT_MODES = ["Cash", "UPI", "Both"]

# Shop details used for printing
SHOP_NAME = "Sri Krishna Agro Centre"
SHOP_PHONE = "6383958656"

SQLITE_DB_PATH = "data/sales_data.db"

def normalize_product_name(name: str) -> str:
    """Normalize product name by stripping and collapsing whitespace."""
    try:
        return ' '.join(str(name or '').strip().split())
    except Exception:
        return str(name or '').strip()

def financial_year_for_date(date: QDate) -> str:
    """Return financial year string for a given QDate."""
    month = date.month()
    year = date.year()
    return f"{year}-{year+1}" if month >= 4 else f"{year-1}-{year}"

def sales_excel_path_month(date: QDate) -> str:
    """Return monthly sales Excel file path for a given QDate."""
    fname = f"Sales_{date.toString('yyyy-MM')}.xlsx"
    return os.path.join(SALES_FILE_DIR, fname)

def sales_excel_path_fy(date: QDate) -> str:
    """Return financial year sales Excel file path for a given QDate."""
    fy = financial_year_for_date(date)
    fname = f"Sales_{fy}.xlsx"
    return os.path.join(SALES_FILE_DIR, fname)

def ensure_sqlite_db() -> None:
    """Ensure the SQLite DB and required tables exist."""
    os.makedirs(SALES_FILE_DIR, exist_ok=True)
    conn = sqlite3.connect(SQLITE_DB_PATH)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS bills (
            bill_number INTEGER PRIMARY KEY,
            date TEXT,
            customer_name TEXT,
            mobile TEXT,
            village TEXT,
            aadhar TEXT,
            product_details TEXT,
            subtotal REAL,
            discount REAL,
            gst_total REAL,
            total REAL,
            payment_mode TEXT,
            cash_amount REAL,
            upi_amount REAL,
            entry_by TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS customers (
            mobile TEXT PRIMARY KEY,
            customer_name TEXT,
            village TEXT,
            aadhar TEXT,
            entry_by TEXT,
            created_at TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS purchase_stock (
            product_name TEXT PRIMARY KEY,
            quantity REAL
        )
    ''')
    # Indices for performance
    try:
        c.execute("CREATE INDEX IF NOT EXISTS idx_bills_date ON bills(date)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_bills_customer ON bills(customer_name)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_customers_mobile ON customers(mobile)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_stock_product ON purchase_stock(product_name)")
    except Exception:
        pass
    conn.commit()
    conn.close()

def get_last_bill_number_from_db() -> int:
    """Get the last bill number from the database."""
    ensure_sqlite_db()
    conn = sqlite3.connect(SQLITE_DB_PATH)
    c = conn.cursor()
    c.execute("SELECT MAX(bill_number) FROM bills")
    row = c.fetchone()
    conn.close()
    return row[0] if row and row[0] else 0

def get_last_bill_number_from_excel(path: str) -> int:
    """Get the last bill number from an Excel file."""
    if not os.path.exists(path):
        return 0
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["Bills"] if "Bills" in wb.sheetnames else wb[wb.sheetnames[0]]
        max_no = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            num = row[0]
            if num is None:
                continue
            try:
                num_int = int(num)
            except Exception:
                try:
                    num_int = int(str(num).strip())
                except Exception:
                    continue
            if num_int > max_no:
                max_no = num_int
        wb.close()
        return max_no
    except Exception:
        return 0

def insert_bill_into_db(bill_data: tuple) -> None:
    """Insert or update a bill record in the database."""
    ensure_sqlite_db()
    conn = sqlite3.connect(SQLITE_DB_PATH)
    c = conn.cursor()
    try:
        c.execute('''
            INSERT INTO bills (
                bill_number, date, customer_name, mobile, village, aadhar,
                product_details, subtotal, discount, gst_total, total,
                payment_mode, cash_amount, upi_amount, entry_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', bill_data)
        conn.commit()
    except sqlite3.IntegrityError:
        c.execute('''
            UPDATE bills SET
            date=?, customer_name=?, mobile=?, village=?, aadhar=?, product_details=?,
            subtotal=?, discount=?, gst_total=?, total=?, payment_mode=?,
            cash_amount=?, upi_amount=?, entry_by=?
            WHERE bill_number=?
        ''', bill_data[1:] + (bill_data[0],))
        conn.commit()
    finally:
        conn.close()

def delete_bill_from_db(bill_number: int) -> None:
    """Delete a bill record from the database."""
    ensure_sqlite_db()
    conn = sqlite3.connect(SQLITE_DB_PATH)
    c = conn.cursor()
    try:
        c.execute('DELETE FROM bills WHERE bill_number = ?', (bill_number,))
        conn.commit()
    finally:
        conn.close()

def insert_or_update_customer_in_db(cust_info: dict) -> None:
    """Insert or update customer info in the database."""
    ensure_sqlite_db()
    conn = sqlite3.connect(SQLITE_DB_PATH)
    c = conn.cursor()
    c.execute('SELECT mobile FROM customers WHERE mobile = ?', (cust_info['mobile'],))
    exists = c.fetchone()
    if exists:
        c.execute('''
            UPDATE customers SET
            customer_name=?, village=?, aadhar=?, entry_by=?, created_at=?
            WHERE mobile=?
        ''', (
            cust_info['cust_name'], cust_info['village'], cust_info['aadhar'],
            cust_info['entry_by'], cust_info['created_at'], cust_info['mobile']
        ))
    else:
        c.execute('''
            INSERT INTO customers (
                mobile, customer_name, village, aadhar, entry_by, created_at
            ) VALUES (?, ?, ?, ?, ?, ?)
        ''', (
            cust_info['mobile'], cust_info['cust_name'], cust_info['village'],
            cust_info['aadhar'], cust_info['entry_by'], cust_info['created_at']
        ))
    conn.commit()
    conn.close()

def reduce_stock_in_db(products: list) -> None:
    """Reduce stock quantities in DB based on products sold."""
    ensure_sqlite_db()
    conn = sqlite3.connect(SQLITE_DB_PATH)
    c = conn.cursor()
    try:
        for p in products:
            pname = normalize_product_name(p.get("Product Name", ""))
            try:
                qty = float(p.get("Quantity", 0))
            except Exception:
                qty = 0.0
            if not pname or qty <= 0:
                continue
            c.execute("SELECT quantity FROM purchase_stock WHERE lower(product_name)=lower(?)", (pname,))
            row = c.fetchone()
            if row is not None:
                try:
                    cur_qty = float(row[0] or 0)
                except Exception:
                    cur_qty = 0.0
                new_qty = max(0.0, cur_qty - qty)
                c.execute("UPDATE purchase_stock SET quantity=? WHERE lower(product_name)=lower(?)", (new_qty, pname))
        conn.commit()
    except Exception as e:
        log_error(f"Stock reduction error: {e}")
    finally:
        conn.close()

def increase_stock_in_db(products: list) -> None:
    """Increase stock quantities in DB (used when editing/deleting bills)."""
    ensure_sqlite_db()
    conn = sqlite3.connect(SQLITE_DB_PATH)
    c = conn.cursor()
    try:
        for p in products:
            pname = normalize_product_name(p.get("Product Name", ""))
            try:
                qty = float(p.get("Quantity", 0))
            except Exception:
                qty = 0.0
            if not pname or qty <= 0:
                continue
            c.execute("SELECT quantity FROM purchase_stock WHERE lower(product_name)=lower(?)", (pname,))
            row = c.fetchone()
            if row is None:
                # If product not in stock table, insert it
                c.execute("INSERT INTO purchase_stock(product_name, quantity) VALUES (?, ?)", (pname, qty))
            else:
                try:
                    cur_qty = float(row[0] or 0)
                except Exception:
                    cur_qty = 0.0
                new_qty = cur_qty + qty
                c.execute("UPDATE purchase_stock SET quantity=? WHERE lower(product_name)=lower(?)", (new_qty, pname))
        conn.commit()
    except Exception as e:
        log_error(f"Stock increase error: {e}")
    finally:
        conn.close()

class SalesWidget(QWidget):
    """Main sales widget for billing and stock management."""
    def __init__(self, auth_manager=None, purchase_excel_path_func=None, parent=None):
        super().__init__(parent)
        self.auth_manager = auth_manager
        self.purchase_excel_path_func = purchase_excel_path_func
        self.current_products = []
        self.last_bill_no = 0
        self.customer_cache = {}
        # Edit mode state
        self.loaded_bill_no = None
        self.loaded_bill_original_products = []
        self.loaded_bill_date_time = None

        self.excel_path_month = sales_excel_path_month(QDate.currentDate())
        self.excel_path_fy = sales_excel_path_fy(QDate.currentDate())
        self.ensure_excel_structure(self.excel_path_month)
        self.ensure_excel_structure(self.excel_path_fy)
        self.user_modified_amounts = False

        if self.purchase_excel_path_func:
            carry_forward_purchase_fy_stock(QDate.currentDate(), self.purchase_excel_path_func)

        db_last = get_last_bill_number_from_db()
        excel_last = max(
            get_last_bill_number_from_excel(self.excel_path_fy),
            get_last_bill_number_from_excel(self.excel_path_month)
        )
        self.last_bill_no = max(db_last, excel_last)

        self.init_ui()
        self.load_purchase_products()
        self.load_customer_cache()

    def ensure_excel_structure(self, path: str) -> None:
        """Ensure Excel file exists with required sheets and headers."""
        os.makedirs(SALES_FILE_DIR, exist_ok=True)
        if not os.path.exists(path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Bills"
            ws.append([
                "Bill Number", "Date", "Customer Name", "Mobile", "Village", "Aadhar",
                "Product Details", "Subtotal", "Discount", "GST Total", "Total",
                "Payment Mode", "Cash Amount", "UPI Amount", "Entry By"
            ])
            ws_cust = wb.create_sheet("CustomerWise")
            ws_cust.append(["Customer Name", "Mobile", "Village", "Aadhar", "Entry By"])
            ws_prod = wb.create_sheet("ProductWise")
            ws_prod.append(["Product Name", "Quantity Sold", "Sale Price", "Bill No", "Date", "Entry By"])
            ws_cat = wb.create_sheet("CategoryWise")
            ws_cat.append([
                "Bill Number", "Date", "Customer Name", "Mobile", "Product Name",
                "Quantity", "Sale Price", "Category", "Entry By"
            ])
            # Category-specific sales sheets
            ws_seeds = wb.create_sheet("SeedsSales")
            ws_seeds.append(["Bill Number", "Date", "Customer Name", "Mobile", "Product Name", "Quantity", "Sale Price", "Entry By"])
            ws_pest = wb.create_sheet("PesticideSales")
            ws_pest.append(["Bill Number", "Date", "Customer Name", "Mobile", "Product Name", "Quantity", "Sale Price", "Entry By"])
            ws_fert = wb.create_sheet("FertilizerSales")
            ws_fert.append(["Bill Number", "Date", "Customer Name", "Mobile", "Product Name", "Quantity", "Sale Price", "Entry By"])
            wb.save(path)

    def init_ui(self) -> None:
        """Initialize the UI components."""
        self.setLayout(QVBoxLayout())

        # --- Customer Details ---
        cust_layout = QHBoxLayout()

        self.cust_name = QLineEdit()
        self.cust_name.setPlaceholderText("Customer Name")
        self.cust_name.setMinimumHeight(36)
        self.cust_name.setStyleSheet("font-size: 18px;")
        self.cust_mobile = QLineEdit()
        self.cust_mobile.setPlaceholderText("Mobile Number")
        self.cust_mobile.setMinimumHeight(36)
        self.cust_mobile.setStyleSheet("font-size: 18px;")
        self.cust_village = QLineEdit()
        self.cust_village.setPlaceholderText("Village")
        self.cust_village.setMinimumHeight(36)
        self.cust_village.setStyleSheet("font-size: 18px;")
        self.cust_aadhar = QLineEdit()
        self.cust_aadhar.setPlaceholderText("Aadhar")
        self.cust_aadhar.setMinimumHeight(36)
        self.cust_aadhar.setStyleSheet("font-size: 18px;")

        # Validators for IDs
        self.cust_mobile.setMaxLength(10)
        self.cust_mobile.setValidator(QRegularExpressionValidator(QRegularExpression(r'^[6-9]\d{9}$'), self))
        self.cust_aadhar.setMaxLength(12)
        self.cust_aadhar.setValidator(QRegularExpressionValidator(QRegularExpression(r'^\d{12}$'), self))

        # On mobile input Enter, fetch customer details
        self.cust_mobile.returnPressed.connect(self.fetch_customer_by_mobile)
        self.cust_name.returnPressed.connect(lambda: self.cust_village.setFocus())
        self.cust_village.returnPressed.connect(lambda: self.cust_aadhar.setFocus())
        self.cust_aadhar.returnPressed.connect(lambda: self.product_name_input.setFocus())

        cust_layout.addWidget(QLabel("Name:"))
        cust_layout.addWidget(self.cust_name)
        cust_layout.addWidget(QLabel("Mobile:"))
        cust_layout.addWidget(self.cust_mobile)
        cust_layout.addWidget(QLabel("Village:"))
        cust_layout.addWidget(self.cust_village)
        cust_layout.addWidget(QLabel("Aadhar:"))
        cust_layout.addWidget(self.cust_aadhar)

        self.layout().addLayout(cust_layout)

        # --- Product Entry ---
        prod_entry_layout = QHBoxLayout()

        self.product_name_input = QLineEdit()
        self.product_name_input.setPlaceholderText("Product Name")
        self.product_name_input.setMinimumHeight(36)
        self.product_name_input.setStyleSheet("font-size: 18px;")
        self.product_qty_input = QLineEdit()
        self.product_qty_input.setPlaceholderText("Quantity")
        self.product_qty_input.setMinimumHeight(36)
        self.product_qty_input.setStyleSheet("font-size: 18px;")
        self.product_price_input = QLineEdit()
        self.product_price_input.setPlaceholderText("Sale Price")
        self.product_price_input.setMinimumHeight(36)
        self.product_price_input.setStyleSheet("font-size: 18px;")

        # Validators for numeric inputs
        num_regex = QRegularExpression(r'^\d+(\.\d+)?$')
        self.product_qty_input.setValidator(QRegularExpressionValidator(num_regex, self))
        self.product_price_input.setValidator(QRegularExpressionValidator(num_regex, self))

        # Setup product autocomplete
        self.product_completer = QCompleter()
        self.product_completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.product_name_input.setCompleter(self.product_completer)
        # Trigger price autofill when product chosen/committed
        try:
            self.product_completer.activated[str].connect(lambda _: self.autofill_price_for_product())
        except Exception:
            pass
        self.product_name_input.editingFinished.connect(self.autofill_price_for_product)
        self.product_name_input.editingFinished.connect(self.update_stock_indicator)
        self.product_qty_input.textChanged.connect(self.update_stock_indicator)

        # On product name enter: autofill price then move focus to qty
        self.product_name_input.returnPressed.connect(self.on_product_name_enter)
        self.product_qty_input.returnPressed.connect(lambda: self.product_price_input.setFocus())
        self.product_price_input.returnPressed.connect(self.add_product_to_list)

        prod_entry_layout.addWidget(QLabel("Product:"))
        prod_entry_layout.addWidget(self.product_name_input)
        prod_entry_layout.addWidget(QLabel("Qty:"))
        prod_entry_layout.addWidget(self.product_qty_input)
        prod_entry_layout.addWidget(QLabel("Price:"))
        prod_entry_layout.addWidget(self.product_price_input)

        # Stock indicator label
        self.stock_label = QLabel("Stock: -")
        self.stock_label.setStyleSheet("color:#555; font-size: 16px;")
        prod_entry_layout.addWidget(self.stock_label)

        self.layout().addLayout(prod_entry_layout)

        # --- Products Table ---
        self.products_table = QTableWidget(0, 4)
        self.products_table.setHorizontalHeaderLabels(["Product Name", "Quantity", "Sale Price", "Amount"])
        self.products_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.products_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.products_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.products_table.setMinimumHeight(120)
        self.products_table.setMaximumHeight(180)  # Make table smaller
        self.products_table.setStyleSheet("font-size: 15px;")
        self.layout().addWidget(self.products_table)

        # Remove selected item button below table
        prod_btn_row = QHBoxLayout()
        remove_btn = QPushButton("Remove Selected Item")
        remove_btn.setMinimumHeight(40)
        remove_btn.setStyleSheet("font-size: 18px;")
        remove_btn.clicked.connect(self.remove_selected_product)
        prod_btn_row.addStretch()
        prod_btn_row.addWidget(remove_btn)
        self.layout().addLayout(prod_btn_row)

        # --- Discount and Payment ---
        discount_layout = QHBoxLayout()
        discount_layout.addWidget(QLabel("Discount:"))
        self.discount_input = QLineEdit("0")
        self.discount_input.setMinimumHeight(36)
        self.discount_input.setStyleSheet("font-size: 18px;")
        discount_layout.addWidget(self.discount_input)

        discount_layout.addWidget(QLabel("Payment Mode:"))
        self.payment_mode_combo = QComboBox()
        self.payment_mode_combo.addItems(PAYMENT_MODES)
        self.payment_mode_combo.setMinimumHeight(36)
        self.payment_mode_combo.setStyleSheet("font-size: 18px;")
        discount_layout.addWidget(self.payment_mode_combo)

        discount_layout.addWidget(QLabel("Cash Amount:"))
        self.cash_amount_input = QLineEdit("0")
        self.cash_amount_input.setMinimumHeight(36)
        self.cash_amount_input.setStyleSheet("font-size: 18px;")
        discount_layout.addWidget(self.cash_amount_input)

        discount_layout.addWidget(QLabel("UPI Amount:"))
        self.upi_amount_input = QLineEdit("0")
        self.upi_amount_input.setMinimumHeight(36)
        self.upi_amount_input.setStyleSheet("font-size: 18px;")
        discount_layout.addWidget(self.upi_amount_input)

        # Track if user changed amounts manually
        self.cash_amount_input.editingFinished.connect(lambda: setattr(self, 'user_modified_amounts', True))
        self.upi_amount_input.editingFinished.connect(lambda: setattr(self, 'user_modified_amounts', True))

        # React to discount changes
        self.discount_input.textChanged.connect(self.maybe_update_payment_amounts)

        # Validators for amounts
        self.discount_input.setValidator(QRegularExpressionValidator(num_regex, self))
        self.cash_amount_input.setValidator(QRegularExpressionValidator(num_regex, self))
        self.upi_amount_input.setValidator(QRegularExpressionValidator(num_regex, self))

        self.layout().addLayout(discount_layout)

        # On payment mode change autofill amounts
        self.payment_mode_combo.currentTextChanged.connect(self.on_payment_mode_changed)

        # --- Total display ---
        total_layout = QHBoxLayout()
        total_layout.addStretch()
        self.total_label = QLabel("Total: ₹0.00")
        self.total_label.setStyleSheet("font-weight: bold; font-size: 22px; color: green;")
        total_layout.addWidget(self.total_label)
        self.layout().addLayout(total_layout)

        # --- Save Button ---
        btn_layout = QHBoxLayout()
        self.save_btn = QPushButton("Save & Print Bill")
        self.save_btn.setMinimumHeight(48)
        self.save_btn.setStyleSheet("font-size: 20px;")
        self.save_btn.clicked.connect(self.save_and_print_bill)
        btn_layout.addStretch()
        btn_layout.addWidget(self.save_btn)
        self.layout().addLayout(btn_layout)

        # --- Bill Summary Section (utilize space below Save & Print) ---
        self.bill_summary_label = QLabel()
        self.bill_summary_label.setStyleSheet("font-size: 16px; color: #333; margin-top: 10px;")
        self.layout().addWidget(self.bill_summary_label)
        self.update_bill_summary()

        # --- Edit/Delete Existing Bill ---
        edit_layout = QHBoxLayout()
        edit_layout.addWidget(QLabel("Bill No:"))
        self.edit_bill_no_input = QLineEdit()
        self.edit_bill_no_input.setPlaceholderText("Enter bill number")
        self.edit_bill_no_input.setMinimumHeight(36)
        self.edit_bill_no_input.setStyleSheet("font-size: 18px;")
        self.edit_bill_no_input.setValidator(QRegularExpressionValidator(QRegularExpression(r'^\d+$'), self))
        edit_layout.addWidget(self.edit_bill_no_input)

        self.load_bill_btn = QPushButton("Load Bill")
        self.load_bill_btn.setMinimumHeight(40)
        self.load_bill_btn.setStyleSheet("font-size: 18px;")
        self.load_bill_btn.clicked.connect(self.load_bill_by_number)
        edit_layout.addWidget(self.load_bill_btn)

        self.update_bill_btn = QPushButton("Update Bill")
        self.update_bill_btn.setMinimumHeight(40)
        self.update_bill_btn.setStyleSheet("font-size: 18px;")
        self.update_bill_btn.clicked.connect(self.update_loaded_bill)
        edit_layout.addWidget(self.update_bill_btn)

        self.delete_bill_btn = QPushButton("Delete Bill")
        self.delete_bill_btn.setMinimumHeight(40)
        self.delete_bill_btn.setStyleSheet("font-size: 18px;")
        self.delete_bill_btn.clicked.connect(self.delete_bill_by_number)
        edit_layout.addWidget(self.delete_bill_btn)

        self.reset_edit_btn = QPushButton("Clear")
        self.reset_edit_btn.setMinimumHeight(40)
        self.reset_edit_btn.setStyleSheet("font-size: 18px;")
        self.reset_edit_btn.clicked.connect(self.reset_all_fields)
        edit_layout.addWidget(self.reset_edit_btn)

        self.layout().addLayout(edit_layout)

    def load_purchase_products(self) -> None:
        """Load product names from DB or Excel for autocomplete."""
        products = set()
        try:
            # Use the same DB file that purchases.py writes to.
            DB_FILE = "purchases.db"
            if os.path.exists(DB_FILE):
                conn = sqlite3.connect(DB_FILE)
                c = conn.cursor()
                c.execute("SELECT DISTINCT product_name FROM purchases")
                for (pname,) in c.fetchall():
                    if pname:
                        products.add(str(pname).strip())
                conn.close()
        except Exception as e:
            print(f"Failed to load products from DB: {e}")

        # Fallback to Excel if DB is empty or unavailable
        if not products and self.purchase_excel_path_func:
            try:
                path = self.purchase_excel_path_func(QDate.currentDate())
                if os.path.exists(path):
                    wb = load_workbook(path, read_only=True, data_only=True)
                    ws = wb["Invoices"] if "Invoices" in wb.sheetnames else wb[wb.sheetnames[0]]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        product_name = row[3] if len(row) > 3 else None
                        if product_name:
                            products.add(str(product_name).strip())
                    wb.close()
            except Exception as e:
                print(f"Failed to load purchase products from Excel: {e}")

        self.purchase_products_cache = sorted(products)
        self.product_model = QStringListModel(self.purchase_products_cache)
        self.product_completer.setModel(self.product_model)
        try:
            self.product_completer.setCompletionMode(QCompleter.PopupCompletion)
        except Exception:
            pass

    def _normalize_mobile(self, s: str) -> str:
        """Normalize mobile number to digits only."""
        return ''.join(ch for ch in s if ch.isdigit())

    def load_customer_cache(self) -> None:
        """Load customer info into cache from DB."""
        self.customer_cache.clear()
        # Load customers into a dict from DB for quick lookup
        try:
            ensure_sqlite_db()
            conn = sqlite3.connect(SQLITE_DB_PATH)
            c = conn.cursor()
            c.execute("SELECT customer_name, mobile, village, aadhar FROM customers")
            for name, mobile, village, aadhar in c.fetchall():
                key = self._normalize_mobile(str(mobile or ""))
                if key:
                    self.customer_cache[key] = dict(
                        cust_name=name or "",
                        village=village or "",
                        aadhar=aadhar or ""
                    )
            conn.close()
        except Exception as e:
            print(f"Failed to load customer cache: {e}")

    def fetch_customer_by_mobile(self) -> None:
        """Fetch customer details by mobile number."""
        raw = self.cust_mobile.text().strip()
        mobile = self._normalize_mobile(raw)
        if not mobile:
            self.cust_name.clear(); self.cust_village.clear(); self.cust_aadhar.clear()
            self.cust_name.setFocus()
            return

        cust = self.customer_cache.get(mobile)
        if not cust:
            # Try DB on-demand
            try:
                ensure_sqlite_db()
                conn = sqlite3.connect(SQLITE_DB_PATH)
                c = conn.cursor()
                c.execute("SELECT customer_name, village, aadhar FROM customers WHERE mobile = ?", (mobile,))
                row = c.fetchone()
                conn.close()
                if row:
                    cust = {"cust_name": row[0] or "", "village": row[1] or "", "aadhar": row[2] or ""}
                    self.customer_cache[mobile] = cust  # cache it
            except Exception as e:
                print(f"Lookup error: {e}")

        if cust:
            self.cust_name.setText(cust['cust_name'])
            self.cust_village.setText(cust['village'])
            self.cust_aadhar.setText(cust['aadhar'])
        else:
            # Optional Excel fallback if you still keep Excel as a source
            try:
                if os.path.exists(CUSTOMER_DATA_FILE):
                    wb = load_workbook(CUSTOMER_DATA_FILE, read_only=True, data_only=True)
                    if "Customers" in wb.sheetnames:
                        ws = wb["Customers"]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            name, mob, village, aadhar, *_ = row
                            if self._normalize_mobile(str(mob or "")) == mobile:
                                self.cust_name.setText(str(name or ""))
                                self.cust_village.setText(str(village or ""))
                                self.cust_aadhar.setText(str(aadhar or ""))
                                break
                    wb.close()
            except Exception as e:
                print(f"Excel fallback lookup error: {e}")

        self.cust_name.setFocus()

    def get_latest_mrp_from_db(self, product_name: str) -> float | None:
        """Get latest MRP for a product from DB."""
        try:
            DB_FILE = "purchases.db"
            if os.path.exists(DB_FILE):
                conn = sqlite3.connect(DB_FILE)
                c = conn.cursor()
                c.execute("SELECT mrp FROM purchases WHERE product_name = ? ORDER BY rowid DESC LIMIT 1", (product_name,))
                row = c.fetchone()
                conn.close()
                if row and row[0] is not None:
                    return float(row[0])
        except Exception as e:
            try:
                log_error(f"MRP lookup DB error: {e}")
            except Exception:
                pass
        return None

    def get_latest_mrp_from_excel(self, product_name: str) -> float | None:
        """Get latest MRP for a product from Excel."""
        if not self.purchase_excel_path_func:
            return None
        try:
            path = self.purchase_excel_path_func(QDate.currentDate())
            if os.path.exists(path):
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb["Invoices"] if "Invoices" in wb.sheetnames else wb[wb.sheetnames[0]]
                last_mrp = None
                for row in ws.iter_rows(min_row=2, values_only=True):
                    pname = row[3] if len(row) > 3 else None
                    mrp = row[6] if len(row) > 6 else None
                    if pname and str(pname).strip().lower() == product_name.strip().lower():
                        try:
                            last_mrp = float(mrp) if mrp is not None else last_mrp
                        except Exception:
                            pass
                wb.close()
                return last_mrp
        except Exception as e:
            try:
                log_error(f"MRP lookup Excel error: {e}")
            except Exception:
                pass
        return None

    def normalize_category(self, category: str) -> str | None:
        """Normalize category string to standard values."""
        if not category:
            return None
        s = str(category).strip().lower()
        if "seed" in s:
            return "Seeds"
        if "pesticid" in s:
            return "Pesticide"
        if "fertil" in s:
            return "Fertilizer"
        return None

    def autofill_price_for_product(self) -> None:
        """Autofill price field for selected product."""
        pname = self.product_name_input.text().strip()
        current_price = self.product_price_input.text().strip()
        if not pname:
            return
        if current_price not in ("", "0", "0.0", "0.00"):
            return
        mrp = self.get_latest_mrp_from_db(pname)
        if mrp is None:
            mrp = self.get_latest_mrp_from_excel(pname)
        if mrp is not None and mrp > 0:
            self.product_price_input.setText(f"{mrp:.2f}")

    def on_product_name_enter(self) -> None:
        """Handle Enter key on product name input."""
        self.autofill_price_for_product()
        self.product_qty_input.setFocus()

    def get_available_stock(self, product_name: str) -> float:
        """Get available stock for a product from DB."""
        try:
            ensure_sqlite_db()
            pname = normalize_product_name(product_name)
            conn = sqlite3.connect(SQLITE_DB_PATH)
            c = conn.cursor()
            c.execute("SELECT quantity FROM purchase_stock WHERE lower(product_name)=lower(?)", (pname,))
            row = c.fetchone()
            conn.close()
            return float(row[0] if row and row[0] is not None else 0.0)
        except Exception:
            return 0.0

    def get_planned_qty_for_product(self, product_name: str) -> float:
        """Get planned quantity for a product in current bill."""
        key = normalize_product_name(product_name).lower()
        try:
            return sum(float(p["Quantity"]) for p in self.current_products if normalize_product_name(p["Product Name"]).lower() == key)
        except Exception:
            return 0.0

    def update_stock_indicator(self) -> None:
        """Update stock indicator label based on product and quantity."""
        pname = self.product_name_input.text().strip()
        if not pname:
            self.stock_label.setText("Stock: -")
            self.stock_label.setStyleSheet("color:#555;")
            return
        available = self.get_available_stock(pname)
        planned = self.get_planned_qty_for_product(pname)
        remaining = max(0.0, available - planned)
        self.stock_label.setText(f"Stock: {remaining:.2f}")
        if remaining <= 0:
            self.stock_label.setStyleSheet("color:red; font-weight:bold;")
        elif remaining <= 5:
            self.stock_label.setStyleSheet("color:orange;")
        else:
            self.stock_label.setStyleSheet("color:green;")

    def open_stock_view(self) -> None:
        """Open a dialog showing available stock."""
        dlg = QDialog(self)
        dlg.setWindowTitle("Available Stock")
        layout = QVBoxLayout(dlg)
        filter_edit = QLineEdit()
        filter_edit.setPlaceholderText("Filter products...")
        layout.addWidget(filter_edit)
        table = QTableWidget(0, 2)
        table.setHorizontalHeaderLabels(["Product Name", "Quantity"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(table)
        def load_data(q: str = ""):
            table.setRowCount(0)
            try:
                ensure_sqlite_db()
                conn = sqlite3.connect(SQLITE_DB_PATH)
                c = conn.cursor()
                if q:
                    like = f"%{q.lower()}%"
                    c.execute("SELECT product_name, quantity FROM purchase_stock WHERE lower(product_name) LIKE ? ORDER BY product_name ASC", (like,))
                else:
                    c.execute("SELECT product_name, quantity FROM purchase_stock ORDER BY product_name ASC")
                for pname, qty in c.fetchall():
                    row = table.rowCount(); table.insertRow(row)
                    table.setItem(row, 0, QTableWidgetItem(str(pname)))
                    try:
                        table.setItem(row, 1, QTableWidgetItem(f"{float(qty or 0):.2f}"))
                    except Exception:
                        table.setItem(row, 1, QTableWidgetItem(str(qty)))
                conn.close()
            except Exception as e:
                QMessageBox.warning(self, "Stock Load Error", f"Failed to load stock: {e}")
        filter_edit.textChanged.connect(load_data)
        load_data()
        dlg.resize(600, 400)
        dlg.exec_()

    def remove_selected_product(self) -> None:
        """Remove selected product from bill."""
        sel = self.products_table.selectionModel().selectedRows()
        if not sel:
            QMessageBox.information(self, "Remove Item", "Select a product to remove.")
            return
        rows = sorted([idx.row() for idx in sel], reverse=True)
        for row in rows:
            if 0 <= row < len(self.current_products):
                try:
                    del self.current_products[row]
                except Exception:
                    pass
            self.products_table.removeRow(row)
        self.update_total_label()
        try:
            self.update_stock_indicator()
        except Exception:
            pass
        self.update_bill_summary()

    def add_product_to_list(self) -> None:
        """Add product entry to bill."""
        pname = self.product_name_input.text().strip()
        try:
            qty = float(self.product_qty_input.text().strip())
            price = float(self.product_price_input.text().strip())
        except ValueError:
            QMessageBox.warning(self, "Input Error", "Quantity and Price must be valid numbers.")
            return

        if not pname:
            QMessageBox.warning(self, "Input Error", "Product name cannot be empty.")
            return
        if qty <= 0 or price <= 0:
            QMessageBox.warning(self, "Input Error", "Quantity and Price must be positive.")
            return

        # Stock check: prevent adding beyond available
        available = self.get_available_stock(pname)
        try:
            planned = sum(p["Quantity"] for p in self.current_products if normalize_product_name(p["Product Name"]).lower() == normalize_product_name(pname).lower())
        except Exception:
            planned = 0.0
        if qty + planned > available + 1e-6:
            remaining = max(0.0, available - planned)
            QMessageBox.warning(self, "Stock Error", f"Only {remaining:.2f} available in stock for {pname}.")
            return

        amount = qty * price
        self.current_products.append({
            "Product Name": pname,
            "Quantity": qty,
            "Sale Price": price,
        })

        row = self.products_table.rowCount()
        self.products_table.insertRow(row)
        self.products_table.setItem(row, 0, QTableWidgetItem(pname))
        self.products_table.setItem(row, 1, QTableWidgetItem(str(qty)))
        self.products_table.setItem(row, 2, QTableWidgetItem(f"{price:.2f}"))
        self.products_table.setItem(row, 3, QTableWidgetItem(f"{amount:.2f}"))

        self.update_total_label()

        # Clear product inputs and set focus back to product name
        self.product_name_input.clear()
        self.product_qty_input.clear()
        self.product_price_input.clear()
        self.product_name_input.setFocus()
        try:
            self.update_stock_indicator()
        except Exception:
            pass
        self.update_bill_summary()

    def update_total_label(self) -> None:
        """Update total label and payment amounts."""
        total = sum(p["Quantity"] * p["Sale Price"] for p in self.current_products)
        discount_text = self.discount_input.text().strip()
        try:
            discount = float(discount_text) if discount_text else 0.0
        except Exception:
            discount = 0.0
        total_after_discount = max(0.0, total - discount)
        self.total_label.setText(f"Total: ₹{total_after_discount:.2f}")
        self.maybe_update_payment_amounts()
        self.update_bill_summary()

    def update_bill_summary(self) -> None:
        """Update bill summary label."""
        num_items = len(self.current_products)
        total_qty = sum(p["Quantity"] for p in self.current_products)
        self.bill_summary_label.setText(
            f"Items in bill: <b>{num_items}</b> &nbsp;&nbsp; Total quantity: <b>{total_qty:.2f}</b>"
        )

    def autofill_payment_amounts(self) -> None:
        """Autofill payment amounts based on payment mode."""
        payment_mode = self.payment_mode_combo.currentText()
        total_text = self.total_label.text().replace("Total: ₹", "")
        try:
            total = float(total_text)
        except Exception:
            total = 0.0
        if payment_mode == "Cash":
            self.cash_amount_input.setText(f"{total:.2f}")
            self.upi_amount_input.setText("0")
        elif payment_mode == "UPI":
            self.upi_amount_input.setText(f"{total:.2f}")
            self.cash_amount_input.setText("0")
        else:  # Both
            self.cash_amount_input.setText("0")
            self.upi_amount_input.setText("0")

    def maybe_update_payment_amounts(self) -> None:
        """Update payment amounts if not manually edited."""
        # Only autofill when user hasn't manually edited the amounts
        if getattr(self, 'user_modified_amounts', False):
            return
        mode = self.payment_mode_combo.currentText()
        if mode in ("Cash", "UPI"):
            self.autofill_payment_amounts()
        elif mode == "Both":
            total_text = self.total_label.text().replace("Total: ₹", "")
            try:
                total = float(total_text)
            except Exception:
                total = 0.0
            cash = self.cash_amount_input.text().strip() or "0"
            upi = self.upi_amount_input.text().strip() or "0"
            try:
                if float(cash) == 0 and float(upi) == 0:
                    self.cash_amount_input.setText(f"{total:.2f}")
                    self.upi_amount_input.setText("0")
            except Exception:
                pass

    def on_payment_mode_changed(self, _) -> None:
        """Handle payment mode change event."""
        # Reset manual edit flag and autofill for single-mode payments
        self.user_modified_amounts = False
        self.autofill_payment_amounts()

    def save_and_print_bill(self) -> None:
        """Save bill, update stock, and show print preview."""
        cust_name = self.cust_name.text().strip()
        mobile = self.cust_mobile.text().strip()
        village = self.cust_village.text().strip()
        aadhar = self.cust_aadhar.text().strip()
        discount_text = self.discount_input.text().strip()
        payment_mode = self.payment_mode_combo.currentText()
        cash_text = self.cash_amount_input.text().strip()
        upi_text = self.upi_amount_input.text().strip()

        if not cust_name or not mobile:
            QMessageBox.warning(self, "Input Error", "Customer Name and Mobile number are required.")
            return
        if not self.current_products:
            QMessageBox.warning(self, "Input Error", "Add at least one product to sale.")
            return
        try:
            discount = float(discount_text)
            if discount < 0:
                raise ValueError
        except ValueError:
            QMessageBox.warning(self, "Input Error", "Discount must be zero or positive number.")
            return
        try:
            cash_amt = float(cash_text)
            upi_amt = float(upi_text)
            if cash_amt < 0 or upi_amt < 0:
                raise ValueError
        except ValueError:
            QMessageBox.warning(self, "Input Error", "Payment amounts must be zero or positive numbers.")
            return

        total_products_value = sum(p["Quantity"] * p["Sale Price"] for p in self.current_products)
        total_payable = max(0.0, total_products_value - discount)
        gst_total = total_products_value * 0.18  # Example GST 18%, adjust if needed

        # Final stock validation across all products
        agg = {}
        for p in self.current_products:
            key = normalize_product_name(p["Product Name"]).lower()
            agg[key] = agg.get(key, 0.0) + float(p["Quantity"])
        insufficient = []
        for key, need in agg.items():
            avail = self.get_available_stock(key)
            if need > avail + 1e-6:
                insufficient.append((key, avail, need))
        if insufficient:
            msg = "\n".join([f"{name}: have {avail:.2f}, need {need:.2f}" for name, avail, need in insufficient])
            QMessageBox.warning(self, "Stock Error", f"Insufficient stock for:\n{msg}")
            return

        if abs((cash_amt + upi_amt) - total_payable) > 0.01:
            QMessageBox.warning(self, "Payment Error", "Sum of payment amounts does not match total payable.")
            return

        # Prepare data for saving
        self.last_bill_no += 1
        bill_date_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        bill_date_str = datetime.now().strftime("%d-%m-%Y")
        products_str = "; ".join(
            f"{p['Product Name']}|{p['Quantity']}|{p['Sale Price']:.2f}"
            for p in self.current_products
        )

        entry_by = self.auth_manager.get_current_user() if self.auth_manager else "Unknown"
        normalized_mobile = self._normalize_mobile(mobile)

        # Save to DB
        bill_data = (
            self.last_bill_no, bill_date_time, cust_name, normalized_mobile, village, aadhar,
            products_str, total_products_value, discount,
            gst_total,  # Example GST 18%, adjust if needed
            total_payable, payment_mode,
            cash_amt, upi_amt, entry_by
        )
        insert_bill_into_db(bill_data)

        # Save/update customer info in DB
        cust_info = {
            "cust_name": cust_name,
            "mobile": normalized_mobile,
            "village": village,
            "aadhar": aadhar,
            "entry_by": entry_by,
            "created_at": bill_date_time,
        }
        insert_or_update_customer_in_db(cust_info)

        # Reduce stock quantities from purchase stock DB
        reduce_stock_in_db(self.current_products)

        # Also write to Sales Excel files (FY and Monthly) and update customer_data.xlsx for cross-module consistency
        try:
            # FY file
            wb_fy = load_workbook(self.excel_path_fy)
            ws_fy = wb_fy["Bills"] if "Bills" in wb_fy.sheetnames else wb_fy[wb_fy.sheetnames[0]]
            ws_fy.append([
                self.last_bill_no, bill_date_str, cust_name, mobile, village, aadhar,
                products_str, total_products_value, discount, gst_total, total_payable,
                payment_mode, cash_amt, upi_amt, entry_by
            ])
            # Per-product details for FY
            ws_prod_fy = wb_fy["ProductWise"] if "ProductWise" in wb_fy.sheetnames else wb_fy.create_sheet("ProductWise")
            ws_cat_fy = wb_fy["CategoryWise"] if "CategoryWise" in wb_fy.sheetnames else wb_fy.create_sheet("CategoryWise")
            # Ensure category sheets exist with headers
            if "SeedsSales" not in wb_fy.sheetnames:
                s = wb_fy.create_sheet("SeedsSales"); s.append(["Bill Number", "Date", "Customer Name", "Mobile", "Product Name", "Quantity", "Sale Price", "Entry By"])
            if "PesticideSales" not in wb_fy.sheetnames:
                s = wb_fy.create_sheet("PesticideSales"); s.append(["Bill Number", "Date", "Customer Name", "Mobile", "Product Name", "Quantity", "Sale Price", "Entry By"])
            if "FertilizerSales" not in wb_fy.sheetnames:
                s = wb_fy.create_sheet("FertilizerSales"); s.append(["Bill Number", "Date", "Customer Name", "Mobile", "Product Name", "Quantity", "Sale Price", "Entry By"])
            for p in self.current_products:
                pname = p["Product Name"]
                qty = p["Quantity"]
                price = p["Sale Price"]
                ws_prod_fy.append([pname, qty, price, self.last_bill_no, bill_date_str, entry_by])
                category_raw = get_product_category_from_db(pname) or ""
                ws_cat_fy.append([
                    self.last_bill_no, bill_date_str, cust_name, mobile, pname, qty, price, category_raw, entry_by
                ])
                cat = self.normalize_category(category_raw)
                if cat == "Seeds":
                    wb_fy["SeedsSales"].append([self.last_bill_no, bill_date_str, cust_name, mobile, pname, qty, price, entry_by])
                elif cat == "Pesticide":
                    wb_fy["PesticideSales"].append([self.last_bill_no, bill_date_str, cust_name, mobile, pname, qty, price, entry_by])
                elif cat == "Fertilizer":
                    wb_fy["FertilizerSales"].append([self.last_bill_no, bill_date_str, cust_name, mobile, pname, qty, price, entry_by])
            wb_fy.save(self.excel_path_fy)

            # Monthly file
            wb_m = load_workbook(self.excel_path_month)
            ws_m = wb_m["Bills"] if "Bills" in wb_m.sheetnames else wb_m[wb_m.sheetnames[0]]
            ws_m.append([
                self.last_bill_no, bill_date_str, cust_name, mobile, village, aadhar,
                products_str, total_products_value, discount, gst_total, total_payable,
                payment_mode, cash_amt, upi_amt, entry_by
            ])
            # Per-product details for Monthly
            ws_prod_m = wb_m["ProductWise"] if "ProductWise" in wb_m.sheetnames else wb_m.create_sheet("ProductWise")
            ws_cat_m = wb_m["CategoryWise"] if "CategoryWise" in wb_m.sheetnames else wb_m.create_sheet("CategoryWise")
            # Ensure category sheets exist with headers
            if "SeedsSales" not in wb_m.sheetnames:
                s = wb_m.create_sheet("SeedsSales"); s.append(["Bill Number", "Date", "Customer Name", "Mobile", "Product Name", "Quantity", "Sale Price", "Entry By"])
            if "PesticideSales" not in wb_m.sheetnames:
                s = wb_m.create_sheet("PesticideSales"); s.append(["Bill Number", "Date", "Customer Name", "Mobile", "Product Name", "Quantity", "Sale Price", "Entry By"])
            if "FertilizerSales" not in wb_m.sheetnames:
                s = wb_m.create_sheet("FertilizerSales"); s.append(["Bill Number", "Date", "Customer Name", "Mobile", "Product Name", "Quantity", "Sale Price", "Entry By"])
            for p in self.current_products:
                pname = p["Product Name"]
                qty = p["Quantity"]
                price = p["Sale Price"]
                ws_prod_m.append([pname, qty, price, self.last_bill_no, bill_date_str, entry_by])
                category_raw = get_product_category_from_db(pname) or ""
                ws_cat_m.append([
                    self.last_bill_no, bill_date_str, cust_name, mobile, pname, qty, price, category_raw, entry_by
                ])
                cat = self.normalize_category(category_raw)
                if cat == "Seeds":
                    wb_m["SeedsSales"].append([self.last_bill_no, bill_date_str, cust_name, mobile, pname, qty, price, entry_by])
                elif cat == "Pesticide":
                    wb_m["PesticideSales"].append([self.last_bill_no, bill_date_str, cust_name, mobile, pname, qty, price, entry_by])
                elif cat == "Fertilizer":
                    wb_m["FertilizerSales"].append([self.last_bill_no, bill_date_str, cust_name, mobile, pname, qty, price, entry_by])
            wb_m.save(self.excel_path_month)
        except Exception as e:
            log_error(f"Failed to write to Sales Excel: {e}")

        # Update centralized customer_data.xlsx for Customers tab
        try:
            update_customer_data_file(
                bill_no=self.last_bill_no,
                date_str=bill_date_str,
                cust_name=cust_name,
                mobile=mobile,
                village=village,
                aadhar=aadhar,
                product_details_str=products_str,
                subtotal=total_products_value,
                discount=discount,
                gst_total=gst_total,
                total=total_payable,
                payment_mode=payment_mode,
                cash_amt=cash_amt,
                upi_amt=upi_amt,
                entry_by=entry_by,
                CUSTOMER_DATA_FILE=CUSTOMER_DATA_FILE,
                ensure_customer_data_file=ensure_customer_data_file,
                add_or_update_customer=add_or_update_customer,
            )
        except Exception as e:
            log_error(f"Failed to update customer_data.xlsx: {e}")

        # Open print preview
        try:
            show_bill_print_preview(
                shop_name=SHOP_NAME,
                shop_phone=SHOP_PHONE,
                bill_no=self.last_bill_no,
                date=bill_date_str,
                customer=cust_name,
                mobile=mobile,
                products=self.current_products,
                subtotal=total_products_value,
                discount=discount,
                gst_total=gst_total,
                total=total_payable,
                parent=self,
            )
        except Exception as e:
            log_error(f"Failed to open print preview: {e}")

        log_info(f"Saved bill #{self.last_bill_no} for {cust_name} total ₹{total_payable:.2f}")
        QMessageBox.information(self, "Success", f"Bill #{self.last_bill_no} saved successfully!")

        # Reset UI for next bill
        self.reset_all_fields()

    def reset_all_fields(self) -> None:
        """Reset all input fields and state for new bill."""
        self.cust_name.clear()
        self.cust_mobile.clear()
        self.cust_village.clear()
        self.cust_aadhar.clear()
        self.discount_input.setText("0")
        self.payment_mode_combo.setCurrentIndex(0)
        self.cash_amount_input.setText("0")
        self.upi_amount_input.setText("0")
        self.user_modified_amounts = False
        self.current_products.clear()
        self.products_table.setRowCount(0)
        self.update_total_label()
        try:
            self.stock_label.setText("Stock: -")
            self.stock_label.setStyleSheet("color:#555;")
        except Exception:
            pass
        self.bill_summary_label.setText("")
        # Clear any edit mode state
        self.loaded_bill_no = None
        self.loaded_bill_original_products = []
        self.loaded_bill_date_time = None

    def parse_product_details_str(self, products_str: str) -> list:
        """Parse product details string into list of dicts."""
        products = []
        if not products_str:
            return products
        for item in [x.strip() for x in str(products_str).split(';') if x.strip()]:
            try:
                name, qty, price = item.split('|')
                qty = float(qty)
                price = float(price)
                products.append({"Product Name": name, "Quantity": qty, "Sale Price": price})
            except Exception:
                continue
        return products

    def populate_products_table_from_list(self, products_list: list) -> None:
        """Populate products table from a list of products."""
        self.products_table.setRowCount(0)
        self.current_products = []
        for p in products_list:
            pname = str(p.get("Product Name", "")).strip()
            try:
                qty = float(p.get("Quantity", 0))
            except Exception:
                qty = 0.0
            try:
                price = float(p.get("Sale Price", 0))
            except Exception:
                price = 0.0
            if not pname:
                continue
            self.current_products.append({"Product Name": pname, "Quantity": qty, "Sale Price": price})
            row = self.products_table.rowCount()
            self.products_table.insertRow(row)
            self.products_table.setItem(row, 0, QTableWidgetItem(pname))
            self.products_table.setItem(row, 1, QTableWidgetItem(f"{qty}"))
            self.products_table.setItem(row, 2, QTableWidgetItem(f"{price:.2f}"))
            self.products_table.setItem(row, 3, QTableWidgetItem(f"{qty*price:.2f}"))
        self.update_total_label()
        self.update_bill_summary()

    def load_bill_by_number(self) -> None:
        """Load bill details by bill number."""
        text = self.edit_bill_no_input.text().strip()
        if not text:
            QMessageBox.information(self, "Load Bill", "Enter a bill number to load.")
            return
        try:
            bill_no = int(text)
        except Exception:
            QMessageBox.warning(self, "Load Bill", "Bill number must be a valid number.")
            return
        ensure_sqlite_db()
        conn = sqlite3.connect(SQLITE_DB_PATH)
        c = conn.cursor()
        c.execute(
            """
            SELECT bill_number, date, customer_name, mobile, village, aadhar,
                   product_details, subtotal, discount, gst_total, total,
                   payment_mode, cash_amount, upi_amount, entry_by
            FROM bills WHERE bill_number = ?
            """,
            (bill_no,)
        )
        row = c.fetchone()
        conn.close()
        if not row:
            QMessageBox.information(self, "Load Bill", f"No bill found for number {bill_no}.")
            return
        # Populate fields
        self.loaded_bill_no = row[0]
        self.loaded_bill_date_time = row[1]
        self.cust_name.setText(str(row[2] or ""))
        self.cust_mobile.setText(str(row[3] or ""))
        self.cust_village.setText(str(row[4] or ""))
        self.cust_aadhar.setText(str(row[5] or ""))
        products_list = self.parse_product_details_str(row[6] or "")
        self.loaded_bill_original_products = [dict(p) for p in products_list]
        self.populate_products_table_from_list(products_list)
        # Discount and payments
        try:
            self.discount_input.setText(f"{float(row[8] or 0):.2f}")
        except Exception:
            self.discount_input.setText("0")
        mode = str(row[11] or "Cash")
        idx = self.payment_mode_combo.findText(mode)
        self.payment_mode_combo.setCurrentIndex(max(0, idx))
        try:
            self.cash_amount_input.setText(f"{float(row[12] or 0):.2f}")
        except Exception:
            self.cash_amount_input.setText("0")
        try:
            self.upi_amount_input.setText(f"{float(row[13] or 0):.2f}")
        except Exception:
            self.upi_amount_input.setText("0")
        self.user_modified_amounts = True  # keep loaded amounts
        QMessageBox.information(self, "Load Bill", f"Bill #{bill_no} loaded. You can now edit and click 'Update Bill'.")

    def update_loaded_bill(self) -> None:
        """Update loaded bill with new details and adjust stock."""
        if self.loaded_bill_no is None:
            QMessageBox.information(self, "Update Bill", "Load a bill first.")
            return
        # Validate inputs similar to save
        cust_name = self.cust_name.text().strip()
        mobile = self.cust_mobile.text().strip()
        village = self.cust_village.text().strip()
        aadhar = self.cust_aadhar.text().strip()
        discount_text = self.discount_input.text().strip() or "0"
        payment_mode = self.payment_mode_combo.currentText()
        cash_text = self.cash_amount_input.text().strip() or "0"
        upi_text = self.upi_amount_input.text().strip() or "0"
        if not cust_name or not mobile:
            QMessageBox.warning(self, "Input Error", "Customer Name and Mobile number are required.")
            return
        if not self.current_products:
            QMessageBox.warning(self, "Input Error", "Add at least one product to sale.")
            return
        try:
            discount = float(discount_text)
            if discount < 0:
                raise ValueError
        except Exception:
            QMessageBox.warning(self, "Input Error", "Discount must be zero or positive number.")
            return
        try:
            cash_amt = float(cash_text)
            upi_amt = float(upi_text)
            if cash_amt < 0 or upi_amt < 0:
                raise ValueError
        except Exception:
            QMessageBox.warning(self, "Input Error", "Payment amounts must be zero or positive numbers.")
            return
        total_products_value = sum(p["Quantity"] * p["Sale Price"] for p in self.current_products)
        total_payable = max(0.0, total_products_value - discount)
        gst_total = total_products_value * 0.18
        if abs((cash_amt + upi_amt) - total_payable) > 0.01:
            QMessageBox.warning(self, "Payment Error", "Sum of payment amounts does not match total payable.")
            return
        # Compute stock deltas
        def to_map(prods):
            m = {}
            names = {}
            for p in prods:
                key = normalize_product_name(p["Product Name"]).lower()
                m[key] = m.get(key, 0.0) + float(p["Quantity"])
                if key not in names:
                    names[key] = p["Product Name"]
            return m, names
        old_map, old_names = to_map(self.loaded_bill_original_products)
        new_map, new_names = to_map(self.current_products)
        keys = set(old_map) | set(new_map)
        # Validate availability for increased quantities
        insufficient = []
        for k in keys:
            delta = new_map.get(k, 0.0) - old_map.get(k, 0.0)
            if delta > 1e-9:
                disp = new_names.get(k) or old_names.get(k) or k
                avail = self.get_available_stock(disp)
                if delta > avail + 1e-6:
                    insufficient.append((disp, avail, delta))
        if insufficient:
            msg = "\n".join([f"{name}: have {avail:.2f}, need extra {need:.2f}" for name, avail, need in insufficient])
            QMessageBox.warning(self, "Stock Error", f"Insufficient stock for updates:\n{msg}")
            return
        # Apply stock adjustments
        inc_list = []  # increase stock (when reduced quantity)
        red_list = []  # reduce stock (when increased quantity)
        for k in keys:
            delta = new_map.get(k, 0.0) - old_map.get(k, 0.0)
            if abs(delta) <= 1e-9:
                continue
            disp = new_names.get(k) or old_names.get(k) or k
            if delta < 0:
                inc_list.append({"Product Name": disp, "Quantity": -delta})
            else:
                red_list.append({"Product Name": disp, "Quantity": delta})
        if inc_list:
            increase_stock_in_db(inc_list)
        if red_list:
            reduce_stock_in_db(red_list)
        # Update DB (keep original date/time)
        products_str = "; ".join(
            f"{p['Product Name']}|{p['Quantity']}|{p['Sale Price']:.2f}" for p in self.current_products
        )
        entry_by = self.auth_manager.get_current_user() if self.auth_manager else "Unknown"
        normalized_mobile = self._normalize_mobile(mobile)
        bill_data = (
            self.loaded_bill_no, self.loaded_bill_date_time, cust_name, normalized_mobile, village, aadhar,
            products_str, total_products_value, discount, gst_total, total_payable,
            payment_mode, cash_amt, upi_amt, entry_by
        )
        insert_bill_into_db(bill_data)
        # Update in-memory originals to new ones
        self.loaded_bill_original_products = [dict(p) for p in self.current_products]
        QMessageBox.information(self, "Update Bill", f"Bill #{self.loaded_bill_no} updated successfully.")

    def delete_bill_by_number(self) -> None:
        """Delete bill by number and restore stock."""
        text = self.edit_bill_no_input.text().strip()
        if not text:
            QMessageBox.information(self, "Delete Bill", "Enter a bill number to delete.")
            return
        try:
            bill_no = int(text)
        except Exception:
            QMessageBox.warning(self, "Delete Bill", "Bill number must be a valid number.")
            return
        ensure_sqlite_db()
        conn = sqlite3.connect(SQLITE_DB_PATH)
        c = conn.cursor()
        c.execute(
            "SELECT product_details FROM bills WHERE bill_number = ?",
            (bill_no,)
        )
        row = c.fetchone()
        conn.close()
        if not row:
            QMessageBox.information(self, "Delete Bill", f"No bill found for number {bill_no}.")
            return
        products_list = self.parse_product_details_str(row[0] or "")
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Delete bill #{bill_no}? This will restore stock quantities.",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        # Restore stock
        if products_list:
            increase_stock_in_db(products_list)
        delete_bill_from_db(bill_no)
        QMessageBox.information(self, "Delete Bill", f"Bill #{bill_no} deleted.")
        # Clear edit mode if we deleted the currently loaded bill
        if self.loaded_bill_no == bill_no:
            self.reset_all_fields()

if __name__ == "__main__":
    import sys
    from PyQt5.QtWidgets import QApplication

    app = QApplication(sys.argv)

    # You can pass your purchase_excel_path_func from purchases.py here
    widget = SalesWidget(purchase_excel_path_func=purchase_excel_path)
    widget.setWindowTitle("Sales Billing System")
    widget.resize(900, 600)
    widget.show()
    sys.exit(app.exec_())
