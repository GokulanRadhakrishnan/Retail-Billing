import os
import sqlite3
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QPushButton,
    QDateEdit, QMessageBox, QTableWidget, QTableWidgetItem, QHeaderView
)
from PyQt5.QtCore import Qt, QDate
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict

DATA_DIR = "data"

def sales_excel_path(date: QDate) -> str:
    """Return sales Excel file path for a given QDate."""
    year = date.year() if date.month() >= 4 else date.year() - 1
    return os.path.join(DATA_DIR, f"Sales_{year}-{year+1}.xlsx")

def purchase_excel_path(date: QDate) -> str:
    """Return purchase Excel file path for a given QDate."""
    year = date.year() if date.month() >= 4 else date.year() - 1
    return os.path.join(DATA_DIR, f"Purchase_{year}-{year+1}.xlsx")

class ReportsWidget(QWidget):
    """Widget for generating sales and inventory reports."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setLayout(QVBoxLayout())
        self.build_ui()

    def build_ui(self) -> None:
        """Build the UI components."""
        filter_layout = QHBoxLayout()

        self.from_date = QDateEdit(QDate.currentDate())
        self.from_date.setDisplayFormat("dd-MM-yyyy")
        filter_layout.addWidget(QLabel("From:"))
        filter_layout.addWidget(self.from_date)

        self.to_date = QDateEdit(QDate.currentDate())
        self.to_date.setDisplayFormat("dd-MM-yyyy")
        filter_layout.addWidget(QLabel("To:"))
        filter_layout.addWidget(self.to_date)

        self.btn_refresh = QPushButton("Refresh Reports")
        self.btn_refresh.clicked.connect(self.refresh_reports)
        filter_layout.addWidget(self.btn_refresh)

        self.layout().addLayout(filter_layout)

        # Matplotlib Figure for charts
        self.fig = Figure(figsize=(10, 6))
        self.canvas = FigureCanvas(self.fig)
        self.layout().addWidget(self.canvas)

        # Tables for Top Products and Customers
        tables_layout = QHBoxLayout()

        self.top_products_table = QTableWidget(0, 3)
        self.top_products_table.setHorizontalHeaderLabels(["Product", "Quantity Sold", "Total Sales (₹)"])
        self.top_products_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tables_layout.addWidget(self.top_products_table)

        self.top_customers_table = QTableWidget(0, 3)
        self.top_customers_table.setHorizontalHeaderLabels(["Customer Name", "Total Purchases", "Number of Bills"])
        self.top_customers_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tables_layout.addWidget(self.top_customers_table)

        self.layout().addLayout(tables_layout)

        # Low stock alert info
        self.low_stock_label = QLabel()
        self.layout().addWidget(self.low_stock_label)

    def refresh_reports(self) -> None:
        """Refresh all report tables and charts."""
        from_date = self.from_date.date().toPyDate()
        to_date = self.to_date.date().toPyDate()
        if from_date > to_date:
            QMessageBox.warning(self, "Input Error", "From date cannot be after To date.")
            return

        # Load sales data for the financial year(s) overlapping the date range
        fy_start_year = from_date.year if from_date.month >= 4 else from_date.year - 1
        fy_end_year = to_date.year if to_date.month >= 4 else to_date.year - 1

        # Aggregate across years if range spans multiple financial years
        sales_data = []
        for yr in range(fy_start_year, fy_end_year + 1):
            path = os.path.join(DATA_DIR, f"Sales_{yr}-{yr+1}.xlsx")
            if os.path.exists(path):
                sales_data.extend(self.load_sales_data(path))

        # Filter data by date range
        filtered_sales = [
            row for row in sales_data if from_date <= datetime.strptime(row['Date'], "%d-%m-%Y").date() <= to_date
        ]

        if not filtered_sales:
            # DB fallback: read from SQLite if Excel had no data
            db_sales = self.load_sales_from_db(from_date, to_date)
            if not db_sales:
                QMessageBox.information(self, "No Data", "No sales data found for the selected date range.")
                self.clear_reports()
                return
            filtered_sales = db_sales

        self.plot_sales_trends(filtered_sales)
        self.populate_top_products(filtered_sales)
        self.populate_top_customers(filtered_sales)
        self.show_low_stock_alerts()

    def load_sales_data(self, path: str) -> list:
        """Load sales data from Excel file."""
        sales = []
        try:
            wb = load_workbook(path, data_only=True)
            sheet = wb["Bills"]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Columns (expected):
                # Bill Number, Date, Customer Name, Mobile, Village, Aadhar,
                # Product Details, Subtotal, Discount, GST Total, Total,
                # Payment Mode, Cash Amount, UPI Amount, Entry By
                sales.append({
                    "Bill Number": row[0],
                    "Date": row[1],
                    "Customer Name": row[2],
                    "Mobile": row[3],
                    "Product Details": row[6],
                    "Subtotal": row[7],
                    "Discount": row[8],
                    "GST Total": row[9],
                    "Total": row[10]
                })
        except Exception as e:
            QMessageBox.warning(self, "Load Error", f"Failed to load sales data: {str(e)}")
        return sales

    def load_sales_from_db(self, from_date, to_date) -> list:
        """Load sales data from SQLite DB."""
        db_path = os.path.join(DATA_DIR, "sales_data.db")
        if not os.path.exists(db_path):
            return []
        rows = []
        try:
            conn = sqlite3.connect(db_path)
            c = conn.cursor()
            # DB stores dates as "%Y-%m-%d %H:%M:%S"; filter by date substring
            c.execute(
                """
                SELECT bill_number, date, customer_name, mobile, product_details,
                       subtotal, discount, gst_total, total
                FROM bills
                WHERE substr(date, 1, 10) >= ? AND substr(date, 1, 10) <= ?
                """,
                (
                    from_date.strftime("%Y-%m-%d"),
                    to_date.strftime("%Y-%m-%d"),
                )
            )
            for (bn, date_str, cname, mobile, prod, subtotal, discount, gst_total, total) in c.fetchall():
                try:
                    # Convert to the same structure used by Excel loader
                    dt_disp = datetime.strptime(date_str[:10], "%Y-%m-%d").strftime("%d-%m-%Y")
                except Exception:
                    dt_disp = date_str
                rows.append({
                    "Bill Number": bn,
                    "Date": dt_disp,
                    "Customer Name": cname,
                    "Mobile": mobile,
                    "Product Details": prod,
                    "Subtotal": subtotal,
                    "Discount": discount,
                    "GST Total": gst_total,
                    "Total": total,
                })
            conn.close()
        except Exception:
            return []
        return rows

    def plot_sales_trends(self, sales: list) -> None:
        """Plot monthly sales trends chart."""
        # Plot monthly sales totals line chart
        monthly_totals = defaultdict(float)
        for sale in sales:
            try:
                date_obj = datetime.strptime(sale['Date'], "%d-%m-%Y")
                month_key = date_obj.strftime("%Y-%m")
                monthly_totals[month_key] += float(sale["Total"] or 0)
            except Exception:
                pass

        if not monthly_totals:
            self.canvas.figure.clear()
            self.canvas.draw()
            return

        months = sorted(monthly_totals.keys())
        totals = [monthly_totals[m] for m in months]

        self.fig.clear()
        ax = self.fig.add_subplot(111)
        ax.plot(months, totals, marker='o', linestyle='-', color='b')
        ax.set_title("Monthly Sales Trend")
        ax.set_xlabel("Month")
        ax.set_ylabel("Sales Total (₹)")
        ax.grid(True)
        ax.tick_params(axis='x', rotation=45)
        self.fig.tight_layout()
        self.canvas.draw()

    def populate_top_products(self, sales: list) -> None:
        """Populate top products table."""
        product_totals = {}
        product_quantities = {}

        for sale in sales:
            product_details = sale["Product Details"]
            if not product_details:
                continue
            # Product details format: "Product1|Qty|Price; Product2|Qty|Price; ..."
            items = [item.strip() for item in product_details.split(';') if item.strip()]
            for item in items:
                try:
                    pname, qty, price = item.split('|')
                    qty = float(qty)
                    price = float(price)
                    product_quantities[pname] = product_quantities.get(pname, 0) + qty
                    product_totals[pname] = product_totals.get(pname, 0) + (qty * price)
                except Exception:
                    continue

        # Sort products by quantity sold descending
        sorted_products = sorted(product_quantities.items(), key=lambda x: x[1], reverse=True)

        self.top_products_table.setRowCount(0)
        for pname, qty in sorted_products[:10]:  # top 10 products
            row_pos = self.top_products_table.rowCount()
            self.top_products_table.insertRow(row_pos)
            self.top_products_table.setItem(row_pos, 0, QTableWidgetItem(pname))
            self.top_products_table.setItem(row_pos, 1, QTableWidgetItem(f"{qty:.2f}"))
            total = product_totals.get(pname, 0)
            self.top_products_table.setItem(row_pos, 2, QTableWidgetItem(f"₹{total:.2f}"))

    def populate_top_customers(self, sales: list) -> None:
        """Populate top customers table."""
        cust_totals = {}
        cust_bills = {}

        for sale in sales:
            cust_name = sale["Customer Name"]
            total = float(sale["Total"] or 0)
            if cust_name:
                cust_totals[cust_name] = cust_totals.get(cust_name, 0) + total
                cust_bills[cust_name] = cust_bills.get(cust_name, 0) + 1

        sorted_customers = sorted(cust_totals.items(), key=lambda x: x[1], reverse=True)

        self.top_customers_table.setRowCount(0)
        for cname, total in sorted_customers[:10]:  # top 10 customers
            row_pos = self.top_customers_table.rowCount()
            self.top_customers_table.insertRow(row_pos)
            self.top_customers_table.setItem(row_pos, 0, QTableWidgetItem(cname))
            self.top_customers_table.setItem(row_pos, 1, QTableWidgetItem(f"₹{total:.2f}"))
            num_bills = cust_bills.get(cname, 0)
            self.top_customers_table.setItem(row_pos, 2, QTableWidgetItem(str(num_bills)))

    def show_low_stock_alerts(self) -> None:
        """Show low stock and near expiry alerts."""
        # Use DB inventory for quantities and Excel for near-expiry detection
        inventory = {}
        db_path = os.path.join(DATA_DIR, "sales_data.db")
        if os.path.exists(db_path):
            try:
                conn = sqlite3.connect(db_path)
                c = conn.cursor()
                c.execute("SELECT product_name, quantity FROM purchase_stock")
                for pname, qty in c.fetchall():
                    inventory[pname] = float(qty or 0)
                conn.close()
            except Exception:
                inventory = {}

        # Near-expiry from Excel
        product_expiries = {}
        path = purchase_excel_path(QDate.currentDate())
        if os.path.exists(path):
            try:
                wb = load_workbook(path, data_only=True)
                sheet = wb["Invoices"]
                now = datetime.now()
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    pname = row[3]
                    expiry_str = row[8]
                    if not pname:
                        continue
                    if expiry_str:
                        try:
                            expiry_date = datetime.strptime(expiry_str, "%d-%m-%Y")
                            if (pname not in product_expiries) or (expiry_date < product_expiries[pname]):
                                product_expiries[pname] = expiry_date
                        except Exception:
                            pass
            except Exception:
                pass

        low_stock_items = [p for p, q in inventory.items() if q <= 5]
        near_expiry_items = [p for p, dt in product_expiries.items() if dt and 0 <= (dt - datetime.now()).days <= 30]

        alert_msgs = []
        if low_stock_items:
            alert_msgs.append(f"Low Stock for: {', '.join(low_stock_items)}")
        if near_expiry_items:
            alert_msgs.append(f"Near Expiry Products: {', '.join(near_expiry_items)}")

        self.low_stock_label.setText("\n".join(alert_msgs) if alert_msgs else "No stock alerts.")

    def clear_reports(self) -> None:
        """Clear all report tables and charts."""
        self.top_products_table.setRowCount(0)
        self.top_customers_table.setRowCount(0)
        self.low_stock_label.setText("")
        self.fig.clear()
        self.canvas.draw()
