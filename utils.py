import os
import re
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Configure a basic logger for the application
LOG_FILE = "app.log"
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.DEBUG,
    format='%(asctime)s:%(levelname)s:%(message)s'
)

# ------------ Financial Year Utilities ------------

FIN_YEAR_START_MONTH = 4  # April

def get_financial_year(date=None):
    """
    Given a date, returns the financial year string in format 'YYYY-YYYY'.
    Financial year starts from April.
    """
    if not date:
        date = datetime.now()
    year = date.year
    if date.month < FIN_YEAR_START_MONTH:
        start = year - 1
        end = year
    else:
        start = year
        end = year + 1
    return f"{start}-{end}"

def get_purchase_excel_filename(date=None):
    fy = get_financial_year(date)
    return f"Purchase_{fy}.xlsx"

def get_sales_excel_filename(date=None):
    fy = get_financial_year(date)
    return f"Sales_{fy}.xlsx"

# ------------ Excel Helper Functions ------------

def ensure_excel_file_with_sheets(path, sheetnames_with_headers):
    """
    Ensure Excel file exists with specified sheets and headers.
    `sheetnames_with_headers` is a dict of sheet_name: list_of_headers
    """
    dir_name = os.path.dirname(path)
    if dir_name and not os.path.exists(dir_name):
        os.makedirs(dir_name)

    if os.path.exists(path):
        try:
            wb = load_workbook(path)
            changed = False
            for sheet_name, headers in sheetnames_with_headers.items():
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)
                    ws = wb[sheet_name]
                    ws.append(headers)
                    changed = True
                else:
                    ws = wb[sheet_name]
                    if ws.max_row == 0 or (ws.max_row == 1 and ws.max_column == 1 and ws['A1'].value is None):
                        ws.append(headers)
                        changed = True
            if changed:
                wb.save(path)
            return wb
        except Exception as e:
            logging.error(f"Failed to load or fix Excel file at {path}: {e}")
            raise
    else:
        try:
            wb = Workbook()
            # remove default sheet
            if 'Sheet' in wb.sheetnames:
                std = wb['Sheet']
                wb.remove(std)
            for sheet_name, headers in sheetnames_with_headers.items():
                ws = wb.create_sheet(sheet_name)
                ws.append(headers)
            wb.save(path)
            return wb
        except Exception as e:
            logging.error(f"Failed to create new Excel file at {path}: {e}")
            raise

# ------------ Validation Functions ------------

def is_valid_mobile_number(mobile):
    """
    Validates Indian mobile numbers (10 digits, starting with 6-9).
    """
    return bool(re.fullmatch(r'[6-9]\d{9}', mobile))

def is_valid_aadhar_number(aadhar):
    """
    Validates Aadhar number (12 digit numeric).
    """
    return bool(re.fullmatch(r'\d{12}', aadhar))

def is_valid_date_string(date_str, date_format="%d-%m-%Y"):
    """
    Validates that a string is a date in the provided format.
    """
    try:
        datetime.strptime(date_str, date_format)
        return True
    except ValueError:
        return False

def parse_date_string(date_str, date_format="%d-%m-%Y"):
    """
    Parses a date string to a datetime object.
    Returns None if invalid.
    """
    try:
        return datetime.strptime(date_str, date_format)
    except ValueError:
        return None

def is_positive_number(value):
    """
    Checks if value can be converted to float and is > 0.
    """
    try:
        return float(value) > 0
    except (ValueError, TypeError):
        return False

def is_non_negative_number(value):
    """
    Checks if value can be converted to float and is >= 0.
    """
    try:
        return float(value) >= 0
    except (ValueError, TypeError):
        return False

# ------------ Autocomplete Helper ------------

def autocomplete_suggestions(word, collection, case_sensitive=False, max_results=10):
    """
    Returns a list of autocomplete suggestions from 'collection' that start with 'word'.
    Respects case sensitivity flag and limits results to 'max_results'.
    """
    if not case_sensitive:
        word = word.lower()
        collection = [w for w in collection if w.lower().startswith(word)]
    else:
        collection = [w for w in collection if w.startswith(word)]

    return collection[:max_results]

# ------------ Error Logging ------------

def log_error(message, exc_info=False):
    """
    Logs an error message with optional stack trace.
    """
    logging.error(message, exc_info=exc_info)

def log_info(message):
    """
    Logs an informational message.
    """
    logging.info(message)

# ------------ Utility Functions ------------

def format_currency(amount, symbol='₹'):
    """
    Format float amount as currency string with Rupee symbol.
    """
    try:
        amt = float(amount)
        return f"{symbol}{amt:,.2f}"
    except (ValueError, TypeError):
        return f"{symbol}0.00"

def safe_float(value, default=0.0):
    """
    Safely convert to float, with fallback default on failure.
    """
    try:
        return float(value)
    except (ValueError, TypeError):
        return default

# ------------ Example usage / test ------------

if __name__ == "__main__":
    # Basic tests
    print("Financial year now:", get_financial_year())
    print("Is valid mobile '9876543210'?", is_valid_mobile_number("9876543210"))
    print("Is valid aadhar '123412341234'?", is_valid_aadhar_number("123412341234"))
    print("Autocomplete 'ure' from ['Urea', 'Urbane', 'Urinal']", autocomplete_suggestions("ure", ["Urea", "Urbane", "Urinal"]))
    print("Format currency 12345.678:", format_currency(12345.678))
    try:
        ensure_excel_file_with_sheets("testdata/testfile.xlsx", {"Sheet1": ["Col1", "Col2"]})
        print("Excel helper works!")
    except Exception as e:
        print("Excel helper error:", e)

CUSTOMER_DATA_FILE = "data/customer_data.xlsx"

def ensure_customer_data_file():
    os.makedirs(os.path.dirname(CUSTOMER_DATA_FILE), exist_ok=True)
    if not os.path.exists(CUSTOMER_DATA_FILE):
        wb = Workbook()
        ws_cust = wb.active
        ws_cust.title = "Customers"
        ws_cust.append(["Customer Name", "Mobile", "Village", "Aadhar", "Entry By", "Created At"])
        ws_ph = wb.create_sheet("PurchaseHistory")
        ws_ph.append([
            "Bill Number", "Date", "Mobile", "Products", "Subtotal", "Discount",
            "GST Total", "Total", "Payment Mode", "Cash Amount", "UPI Amount", "Entry By"
        ])
        wb.save(CUSTOMER_DATA_FILE)

def add_or_update_customer(
    cust_name: str,
    mobile: str,
    village: str,
    aadhar: str,
    entry_by: str,
    created_at: str,
    ws_cust,
):
    header = [cell.value for cell in ws_cust[1]]
    idx_name = header.index("Customer Name") + 1
    idx_mobile = header.index("Mobile") + 1
    idx_village = header.index("Village") + 1
    idx_aadhar = header.index("Aadhar") + 1
    idx_entry_by = header.index("Entry By") + 1
    idx_created_at = header.index("Created At") + 1

    found_row = None
    for row in ws_cust.iter_rows(min_row=2):
        if str(row[idx_mobile - 1].value) == str(mobile):
            found_row = row[0].row
            break

    if found_row:
        if ws_cust.cell(row=found_row, column=idx_name).value in [None, ""] and cust_name:
            ws_cust.cell(row=found_row, column=idx_name, value=cust_name)
        if ws_cust.cell(row=found_row, column=idx_village).value in [None, ""] and village:
            ws_cust.cell(row=found_row, column=idx_village, value=village)
        if ws_cust.cell(row=found_row, column=idx_aadhar).value in [None, ""] and aadhar:
            ws_cust.cell(row=found_row, column=idx_aadhar, value=aadhar)
        if entry_by:
            ws_cust.cell(row=found_row, column=idx_entry_by, value=entry_by)
        ws_cust.cell(row=found_row, column=idx_created_at, value=created_at)
    else:
        ws_cust.append([cust_name, mobile, village, aadhar, entry_by, created_at])
from datetime import datetime
from openpyxl import load_workbook

def update_customer_data_file(
    bill_no: int,
    date_str: str,
    cust_name: str,
    mobile: str,
    village: str,
    aadhar: str,
    product_details_str: str,
    subtotal: float,
    discount: float,
    gst_total: float,
    total: float,
    payment_mode: str,
    cash_amt: float,
    upi_amt: float,
    entry_by: str,
    CUSTOMER_DATA_FILE: str,
    ensure_customer_data_file,
    add_or_update_customer,
):
    """
    Updates customer_data.xlsx with customer info and purchase history.
    Auto-fills missing customer fields if possible.
    Removes purchase history older than 3 years.

    Parameters:
    - bill_no: Bill number (int)
    - date_str: Date string in format 'dd-mm-yyyy' (str)
    - cust_name: Customer Name (str)
    - mobile: Customer Mobile Number (str)
    - village: Customer Village (str)
    - aadhar: Customer Aadhar (str)
    - product_details_str: Products in bill as string (str)
    - subtotal: Subtotal before discount and GST (float)
    - discount: Discount applied (float)
    - gst_total: Total GST amount (float)
    - total: Final total amount (float)
    - payment_mode: Payment mode string ('Cash', 'UPI', 'Both') (str)
    - cash_amt: Cash amount paid (float)
    - upi_amt: UPI amount paid (float)
    - entry_by: Username of staff entering bill (str)
    - CUSTOMER_DATA_FILE: Path to customer_data.xlsx (str)
    - ensure_customer_data_file: Function to ensure customer data file exists (callable)
    - add_or_update_customer: Function to add or update customer details in Excel (callable)
    """

    # Ensure file and sheets exist
    ensure_customer_data_file()

    # Load workbook and sheets
    wb = load_workbook(CUSTOMER_DATA_FILE)
    ws_cust = wb["Customers"]
    ws_ph = wb["PurchaseHistory"]
    now_str = datetime.now().strftime("%Y-%m-%d")

    # Update or insert customer info
    add_or_update_customer(
        cust_name=cust_name,
        mobile=mobile,
        village=village,
        aadhar=aadhar,
        entry_by=entry_by,
        created_at=now_str,
        ws_cust=ws_cust
    )

    # Append new purchase history record
    ws_ph.append([
        bill_no, date_str, mobile, product_details_str,
        subtotal, discount, gst_total, total,
        payment_mode, cash_amt, upi_amt, entry_by
    ])

    # Remove purchase history older than 3 years
    from datetime import timedelta
    cutoff_date = datetime.now() - timedelta(days=3*365)
    rows_to_delete = []
    for row in ws_ph.iter_rows(min_row=2):
        date_cell = row[1].value  # Date is second column in PurchaseHistory sheet
        try:
            dt = datetime.strptime(date_cell, "%d-%m-%Y")
            if dt < cutoff_date:
                rows_to_delete.append(row[0].row)
        except Exception:
            # If date invalid or missing, skip deletion for safety
            continue

    for r in reversed(rows_to_delete):
        ws_ph.delete_rows(r)

    # Save workbook
    wb.save(CUSTOMER_DATA_FILE)

