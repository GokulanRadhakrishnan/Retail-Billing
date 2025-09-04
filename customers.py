import os
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
    QTextEdit, QMessageBox, QGroupBox, QTableWidget, QTableWidgetItem, QHeaderView, QComboBox
)
from PyQt5.QtCore import Qt, QDate
from openpyxl import load_workbook, Workbook
from datetime import datetime
from utils import update_customer_data_file, ensure_customer_data_file, add_or_update_customer, CUSTOMER_DATA_FILE

CUSTOMER_FILE_DIR = "data"
LOYALTY_PER_RS = 100  # 1 point per 100Rs spent

def sales_excel_path(date: QDate):
    year = date.year() if date.month() >= 4 else date.year() - 1
    return os.path.join(CUSTOMER_FILE_DIR, f"Sales_{year}-{year+1}.xlsx")

class CustomerWidget(QWidget):
    def __init__(self, auth_manager=None, parent=None):
        super().__init__(parent)
        self.auth_manager = auth_manager
        self.setLayout(QVBoxLayout())
        self.current_mobile = ""
        self.sales_file = sales_excel_path(QDate.currentDate())
        self.build_ui()

    def build_ui(self):
        form_group = QGroupBox("Customer Lookup")
        form_layout = QHBoxLayout()
        form_group.setLayout(form_layout)

        self.mobile_input = QLineEdit()
        self.mobile_input.setPlaceholderText("Enter Mobile Number")
        form_layout.addWidget(QLabel("Mobile:"))
        form_layout.addWidget(self.mobile_input)
        btn_search = QPushButton("Search")
        btn_search.clicked.connect(self.handle_search)
        form_layout.addWidget(btn_search)

        btn_clear = QPushButton("Clear")
        btn_clear.clicked.connect(self.clear_fields)
        form_layout.addWidget(btn_clear)
        self.layout().addWidget(form_group)

        # Customer Details Display
        self.details_text = QTextEdit()
        self.details_text.setReadOnly(True)
        self.layout().addWidget(self.details_text)

        # Loyalty controls
        loyalty_group = QGroupBox("Loyalty Points Management")
        loyalty_layout = QHBoxLayout()
        loyalty_group.setLayout(loyalty_layout)
        self.points_label = QLabel("Loyalty Points: 0")
        loyalty_layout.addWidget(self.points_label)
        self.points_input = QLineEdit("0")
        self.points_input.setFixedWidth(80)
        loyalty_layout.addWidget(QLabel("Adjust Points:"))
        loyalty_layout.addWidget(self.points_input)
        self.adjust_reason = QLineEdit()
        self.adjust_reason.setPlaceholderText("Reason/Reference")
        loyalty_layout.addWidget(self.adjust_reason)
        btn_add = QPushButton("Add")
        btn_add.clicked.connect(lambda: self.adjust_points("add"))
        loyalty_layout.addWidget(btn_add)
        btn_redeem = QPushButton("Redeem")
        btn_redeem.clicked.connect(lambda: self.adjust_points("redeem"))
        loyalty_layout.addWidget(btn_redeem)
        self.layout().addWidget(loyalty_group)

        # Purchase History Table
        hist_group = QGroupBox("Purchase History")
        hist_layout = QVBoxLayout()
        hist_group.setLayout(hist_layout)
        self.hist_table = QTableWidget(0, 5)
        self.hist_table.setHorizontalHeaderLabels(["Bill #", "Date", "Products", "Amount", "Discount"])
        self.hist_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        hist_layout.addWidget(self.hist_table)
        self.layout().addWidget(hist_group)

    def handle_search(self):
        mobile = self.mobile_input.text().strip()
        if not mobile:
            QMessageBox.warning(self, "Input Error", "Enter a mobile number.")
            return
        self.current_mobile = mobile
        self.sales_file = sales_excel_path(QDate.currentDate())
        self.refresh_customer_info()
        self.refresh_purchase_history()

    def clear_fields(self):
        self.mobile_input.clear()
        self.details_text.clear()
        self.points_label.setText("Loyalty Points: 0")
        self.points_input.setText("0")
        self.adjust_reason.clear()
        self.hist_table.setRowCount(0)
        self.current_mobile = ""

    def refresh_customer_info(self):
        if not os.path.exists(CUSTOMER_DATA_FILE):
            self.details_text.setText("Customer data file missing.")
            self.points_label.setText("Loyalty Points: 0")
            return
        try:
            wb = load_workbook(CUSTOMER_DATA_FILE)
            ws = wb["Customers"]
            found = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                name, mobile, village, aadhar, *_ = row
                if str(mobile) == self.current_mobile:
                    info = f"Name: {name}\nMobile: {mobile}\nVillage: {village}\nAadhar: {aadhar}"
                    self.details_text.setText(info)
                    found = True
                    break
            if not found:
               self.details_text.setText("Customer not found in records.")
               self.points_label.setText("Loyalty Points: 0")
               return
            points = self.read_loyalty_points()
            self.points_label.setText(f"Loyalty Points: {points}")
        except Exception as e:
            self.details_text.setText(f"Error loading customer info: {e}")
            self.points_label.setText("Loyalty Points: 0")
            
    def refresh_purchase_history(self):
        self.hist_table.setRowCount(0)
        if not os.path.exists(CUSTOMER_DATA_FILE):
           return
        try:
            wb = load_workbook(CUSTOMER_DATA_FILE)
            ws = wb["PurchaseHistory"]
            from datetime import datetime, timedelta
            cutoff_date = datetime.now() - timedelta(days=3*365)
            count = 0
            total_value = 0
            total_discount = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                bill_no, date_str, mobile, prod_str, subtotal, discount, _, total, *_ = row
                if str(mobile) == self.current_mobile:
                   try:
                       dt = datetime.strptime(date_str, "%d-%m-%Y")
                       if dt < cutoff_date:
                          continue  # Skip older records
                   except:
                     continue
                   rowpos = self.hist_table.rowCount()
                   self.hist_table.insertRow(rowpos)
                   self.hist_table.setItem(rowpos, 0, QTableWidgetItem(str(bill_no)))
                   self.hist_table.setItem(rowpos, 1, QTableWidgetItem(date_str))
                   self.hist_table.setItem(rowpos, 2, QTableWidgetItem(str(prod_str)))
                   self.hist_table.setItem(rowpos, 3, QTableWidgetItem(f"{total:.2f}" if total else ""))
                   self.hist_table.setItem(rowpos, 4, QTableWidgetItem(f"{discount:.2f}" if discount else ""))
                   total_value += total if total else 0
                   total_discount += discount if discount else 0
                   count += 1
            self.details_text.append(f"\nTotal Purchases: {count}\nTotal Value: ₹{total_value:.2f}\nTotal Discount: ₹{total_discount:.2f}")
        except Exception:
          pass

    def read_loyalty_points(self):
        # Optionally, store a loyalty_points.xlsx per financial year.
        points_file = os.path.join(CUSTOMER_FILE_DIR, "loyalty_points.xlsx")
        if not os.path.exists(points_file):
            return 0
        try:
            wb = load_workbook(points_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                mobile, pts = row[0], row[1]
                if str(mobile) == self.current_mobile:
                    return pts
        except Exception:
            pass
        return 0

    def write_loyalty_points(self, new_points, reason):
        points_file = os.path.join(CUSTOMER_FILE_DIR, "loyalty_points.xlsx")
        os.makedirs(CUSTOMER_FILE_DIR, exist_ok=True)
        if os.path.exists(points_file):
            wb = load_workbook(points_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Mobile", "Points", "LastUpdated", "Reason", "StaffUser"])
        mob_found = False
        # Update if present
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == self.current_mobile:
                ws.cell(row=row[0].row, column=2, value=new_points)
                ws.cell(row=row[0].row, column=3, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
                ws.cell(row=row[0].row, column=4, value=reason)
                ws.cell(row=row[0].row, column=5, value=self.auth_manager.get_current_user() if self.auth_manager else "unknown")
                mob_found = True
                break
        if not mob_found:
            ws.append([
                self.current_mobile, new_points, datetime.now().strftime("%Y-%m-%d %H:%M"),
                reason, self.auth_manager.get_current_user() if self.auth_manager else "unknown"
            ])
        wb.save(points_file)

    def adjust_points(self, mode):
        if not self.current_mobile:
            QMessageBox.warning(self, "No Customer", "Enter and search for a customer first.")
            return
        # Permission enforcement for adjust/redeem (admin-only for redeem):
        if mode == "redeem" and (not self.auth_manager or not self.auth_manager.has_role("admin")):
            QMessageBox.warning(self, "Permission Denied", "Only admin may redeem points!")
            return
        pts_str = self.points_input.text().strip()
        try:
            pts = int(pts_str)
            if pts <= 0:
                raise ValueError
        except Exception:
            QMessageBox.warning(self, "Input Error", "Enter a valid positive integer for points.")
            return
        cur_points = self.read_loyalty_points()
        if mode == "add":
            new_points = cur_points + pts
        elif mode == "redeem":
            if pts > cur_points:
                QMessageBox.warning(self, "Not Enough Points", f"Cannot redeem {pts} points; only {cur_points} available.")
                return
            new_points = cur_points - pts
        else:
            return
        reason = self.adjust_reason.text().strip()
        if not reason:
            reason = f"{'Added' if mode == 'add' else 'Redeemed'} points"
        self.write_loyalty_points(new_points, reason)
        self.points_label.setText(f"Loyalty Points: {new_points}")
        QMessageBox.information(self, "Points Updated", f"Customer now has {new_points} loyalty points.")

